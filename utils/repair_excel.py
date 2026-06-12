"""
repair_excel.py — 손상된 ceria_synthesis_database.xlsx 복구

복구 전략 (순서대로 시도):
  1. openpyxl read_only=False 로 직접 열기 → 성공 시 재저장
  2. zipfile 수준에서 XML 추출 → 데이터 행 직접 파싱
  3. 모두 실패 시 ceria_samples_merged.csv 에서 재구성

실행:  python repair_excel.py
"""
import os, sys, zipfile, io, re
import pandas as pd
import numpy as np

BASE   = r"d:\머신러닝 교육\ceria_pipeline_data"
OUTPUT = os.path.join(BASE, "output")
XLSX   = os.path.join(OUTPUT, "ceria_synthesis_database.xlsx")
XLSX_BAK = XLSX + ".bak"
SAMPLES  = os.path.join(OUTPUT, "ceria_samples_merged.csv")

print("=" * 60)
print("Excel 복구 시작")
print(f"대상: {XLSX}")
print("=" * 60)


# ── 전략 1: openpyxl 일반 모드 (read_only=False) ─────────────────────────────
def try_openpyxl():
    print("\n[전략1] openpyxl read_only=False …")
    try:
        from openpyxl import load_workbook
        wb = load_workbook(XLSX, read_only=False, data_only=True)
        print(f"  시트 목록: {wb.sheetnames}")
        ws = wb.active
        data = list(ws.values)
        if not data:
            print("  데이터 없음")
            return None
        headers = [str(c) if c is not None else "" for c in data[0]]
        rows    = [[c for c in row] for row in data[1:]]
        df = pd.DataFrame(rows, columns=headers)
        print(f"  로드 성공: {len(df):,}행 × {len(df.columns)}열")
        return wb, df
    except Exception as e:
        print(f"  실패: {e}")
        return None


# ── 전략 2: pandas engine_kwargs ─────────────────────────────────────────────
def try_pandas_kwargs():
    print("\n[전략2] pandas engine_kwargs read_only=False …")
    try:
        df = pd.read_excel(XLSX, sheet_name=0,
                           engine="openpyxl",
                           engine_kwargs={"read_only": False, "data_only": True})
        print(f"  로드 성공: {len(df):,}행 × {len(df.columns)}열")
        return df
    except Exception as e:
        print(f"  실패: {e}")
        return None


# ── 전략 3: zipfile 수준 XML 파싱 ────────────────────────────────────────────
def try_zip_repair():
    print("\n[전략3] ZIP 내부 XML 직접 파싱 …")
    try:
        zf = zipfile.ZipFile(XLSX, 'r')
        names = zf.namelist()
        print(f"  ZIP 항목 수: {len(names)}")

        # shared strings
        shared = []
        if "xl/sharedStrings.xml" in names:
            raw = zf.read("xl/sharedStrings.xml").decode("utf-8", errors="ignore")
            shared = re.findall(r"<t(?:\s[^>]*)?>([^<]*)</t>", raw)

        # sheet1 XML
        sheet_names = [n for n in names if re.match(r"xl/worksheets/sheet\d+\.xml", n)]
        if not sheet_names:
            print("  시트 XML 없음")
            return None
        sheet_xml = zf.read(sheet_names[0]).decode("utf-8", errors="ignore")
        zf.close()

        # 행/셀 파싱
        rows = []
        for row_m in re.finditer(r"<row[^>]*>(.*?)</row>", sheet_xml, re.S):
            cells = []
            for cell_m in re.finditer(r'<c r="([A-Z]+\d+)"(?:[^>]*t="([^"]*)")?[^>]*><v>([^<]*)</v>', row_m.group(1)):
                col_ref, typ, val = cell_m.groups()
                if typ == "s":
                    idx = int(val)
                    val = shared[idx] if idx < len(shared) else ""
                cells.append((col_ref, val))
            rows.append(cells)

        if not rows:
            print("  행 없음")
            return None

        # 컬럼 알파벳 → 인덱스
        def col_to_idx(col_str):
            n = 0
            for c in col_str:
                n = n * 26 + (ord(c) - ord('A') + 1)
            return n - 1

        max_col = max(col_to_idx(re.sub(r'\d', '', c)) for row in rows for c, _ in row)
        matrix = []
        for row in rows:
            r = [""] * (max_col + 1)
            for cell_ref, val in row:
                ci = col_to_idx(re.sub(r'\d', '', cell_ref))
                r[ci] = val
            matrix.append(r)

        headers = matrix[0]
        df = pd.DataFrame(matrix[1:], columns=headers)
        print(f"  파싱 성공: {len(df):,}행 × {len(df.columns)}열")
        return df

    except Exception as e:
        print(f"  실패: {e}")
        return None


# ── 전략 4: ceria_samples_merged.csv 에서 재구성 ─────────────────────────────
def rebuild_from_csv():
    print("\n[전략4] ceria_samples_merged.csv 에서 재구성 …")
    if not os.path.exists(SAMPLES):
        print(f"  {SAMPLES} 없음 — 재구성 불가")
        return None
    df_s = pd.read_csv(SAMPLES, dtype=str, low_memory=False)
    print(f"  샘플 CSV: {len(df_s):,}행")

    # 논문별 첫 번째 샘플 값으로 집약 (paper_id 또는 doi 기준)
    id_col = "doi" if "doi" in df_s.columns else "paper_id"
    agg = df_s.groupby(id_col).first().reset_index()
    print(f"  논문 수: {len(agg):,}편")
    print("  ※ 논문 메타데이터(title/authors/year 등)는 손실될 수 있음")
    return agg


# ── 메인 실행 ─────────────────────────────────────────────────────────────────
result = try_openpyxl()
df_recovered = None

MIN_PAPERS = 4000  # 이 수 이상이어야 완전 복구로 간주
REQUIRED_COLS = {"doi", "synthesis_method"}  # 평탄화된 구조 확인용


def _is_complete(df_check):
    return len(df_check) >= MIN_PAPERS and REQUIRED_COLS.issubset(set(df_check.columns))


if result:
    wb_obj, df_recovered = result
    if _is_complete(df_recovered):
        try:
            with pd.ExcelWriter(XLSX, engine="openpyxl") as writer:
                df_recovered.to_excel(writer, sheet_name="합성조건", index=False)
            print(f"\n✓ 복구 완료 (전략1): {XLSX}")
            sys.exit(0)
        except Exception as e:
            print(f"  재저장 실패: {e}")
    else:
        missing = REQUIRED_COLS - set(df_recovered.columns)
        print(f"  전략1: {len(df_recovered):,}편 / 누락 컬럼 {missing or '없음'} — 전략5로 진행")

if df_recovered is None or not _is_complete(df_recovered):
    r2 = try_pandas_kwargs()
    if r2 is not None and _is_complete(r2):
        try:
            with pd.ExcelWriter(XLSX, engine="openpyxl") as writer:
                r2.to_excel(writer, sheet_name="합성조건", index=False)
            print(f"\n✓ 복구 완료 (전략2): {XLSX}")
            sys.exit(0)
        except Exception as e:
            print(f"  재저장 실패: {e}")
    elif r2 is not None:
        print(f"  전략2: {len(r2):,}편 — 불완전, 전략5로 진행")
        if df_recovered is None or len(r2) > len(df_recovered):
            df_recovered = r2

if df_recovered is None or not _is_complete(df_recovered):
    r3 = try_zip_repair()
    if r3 is not None and _is_complete(r3):
        try:
            with pd.ExcelWriter(XLSX, engine="openpyxl") as writer:
                r3.to_excel(writer, sheet_name="합성조건", index=False)
            print(f"\n✓ 복구 완료 (전략3): {XLSX}")
            sys.exit(0)
        except Exception as e:
            print(f"  재저장 실패: {e}")
    elif r3 is not None:
        print(f"  전략3: {len(r3):,}편 — 불완전, 전략5로 진행")
        if df_recovered is None or len(r3) > len(df_recovered):
            df_recovered = r3

# ── 전략 5 (권장): ceria_dataset_full.jsonl 에서 완전 복구 ───────────────────
def _flatten_jsonl(rec):
    """JSONL 중첩 레코드를 Excel 평탄 컬럼 구조로 변환."""
    row = {
        "paper_id":         rec.get("id"),
        "title":            rec.get("title"),
        "year":             rec.get("year"),
        "journal":          rec.get("journal"),
        "doi":              rec.get("doi"),
        "citation_count":   rec.get("citation_count"),
        "source_api":       rec.get("source"),
        "completeness_score": rec.get("completeness_score"),
        "is_synthesis_paper": rec.get("is_synthesis_paper"),
    }
    sc = rec.get("synthesis_conditions")
    if isinstance(sc, dict):
        row.update(sc)
    return row


def rebuild_from_jsonl(df_partial=None):
    print("\n[전략5] ceria_dataset_full.jsonl 에서 완전 복구 (4,388편) …")
    jsonl_path = os.path.join(OUTPUT, "ceria_dataset_full.jsonl")
    if not os.path.exists(jsonl_path):
        print(f"  {jsonl_path} 없음")
        return None
    import json
    raw = []
    with open(jsonl_path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                raw.append(json.loads(line))
    if not raw:
        print("  데이터 없음")
        return None
    # 중첩 → 평탄화
    df_jsonl = pd.DataFrame([_flatten_jsonl(r) for r in raw])
    print(f"  JSONL 로드+평탄화: {len(df_jsonl):,}편 × {len(df_jsonl.columns)}열")

    # df_partial이 synthesis_method를 가진 정상 구조일 때만 병합
    partial_ok = (df_partial is not None and len(df_partial) > 100
                  and "synthesis_method" in df_partial.columns)
    if partial_ok:
        print(f"  기존 데이터({len(df_partial):,}편 × {len(df_partial.columns)}열)와 병합 …")
        all_cols = list(dict.fromkeys(list(df_partial.columns) + list(df_jsonl.columns)))
        existing_dois = set(df_partial["doi"].dropna().astype(str).str.strip())
        extra = df_jsonl[~df_jsonl["doi"].astype(str).str.strip().isin(existing_dois)].copy()
        df = pd.concat(
            [df_partial.reindex(columns=all_cols),
             extra.reindex(columns=all_cols)],
            ignore_index=True
        )
        print(f"  병합 후: {len(df):,}편 × {len(df.columns)}열")
    else:
        df = df_jsonl
        print(f"  기존 구조 불완전 — JSONL 평탄화 38열 직접 사용 ({len(df):,}편)")

    # ceria_samples_merged.csv 에서 ce_precursor / solvent 역전파
    if os.path.exists(SAMPLES):
        df_s = pd.read_csv(SAMPLES, dtype=str, low_memory=False)
        _bad = {"", "nan", "none", "null", "n/a", "na"}
        for col in ["ce_precursor", "solvent"]:
            if col not in df_s.columns:
                continue
            valid = df_s[df_s["doi"].notna() & df_s[col].notna() &
                         ~df_s[col].str.strip().str.lower().isin(_bad)]
            best = valid.groupby("doi")[col].first()
            # 컬럼이 없거나 float64면 object로 강제 변환
            if col not in df.columns:
                df[col] = None
            df[col] = df[col].astype(object)
            # 빈 셀 마스크
            null_mask = df[col].isna() | (
                df[col].astype(str).str.strip().str.lower().isin(_bad)
            )
            # 벡터화: doi → best 값 매핑
            mapped = df["doi"].map(best)
            fill_mask = null_mask & mapped.notna()
            df.loc[fill_mask, col] = mapped[fill_mask]
            print(f"  CSV 역전파 [{col}]: {null_mask.sum()}개 빈 셀 중 {int(fill_mask.sum())}개 채움")

    return df


# ── 전략 5 실행 ───────────────────────────────────────────────────────────────
if df_recovered is None or not _is_complete(df_recovered):
    df_full = rebuild_from_jsonl(df_recovered)
    if df_full is not None and len(df_full) > 4000:
        try:
            tmp = XLSX.replace(".xlsx", "_tmp.xlsx")
            df_full.to_excel(tmp, index=False, sheet_name="합성조건")
            os.replace(tmp, XLSX)
            print(f"\n✓ 완전 복구 완료 (전략5 — JSONL 기반): {XLSX}")
            print(f"  총 {len(df_full):,}편 복구됨")
            print("  fill_keywords.py 재실행 권장")
            sys.exit(0)
        except Exception as e:
            print(f"  재저장 실패: {e}")

if df_recovered is not None:
    print(f"\n⚠ 부분 복구 완료 (전략4 — CSV 기반, {len(df_recovered):,}편)")
    print("  pipeline.py 재실행으로 4,388편 완전 복구 필요")
    sys.exit(0)

print("\n✗ 모든 복구 전략 실패.")
print("  pipeline.py 재실행으로 Excel 재생성 필요.")
sys.exit(1)
