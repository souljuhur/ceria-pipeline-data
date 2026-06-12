"""
format_excel.py — ceria_synthesis_database.xlsx 서식 개선

적용 내용:
  1. 상단 요약 블록 (총 논문수, ML 가능 수, 전문 보유 수, composite 커버리지)
  2. 컬럼 헤더 강조 (파란 배경, 흰 글씨, 굵게)
  3. 모든 데이터 셀 테두리 적용 (얇은 회색 선)
  4. 홀짝 행 교대 배경색
  5. 열 너비 자동 조정
  6. 틀 고정 (헤더 행)
  7. 자동 필터

실행: python format_excel.py
"""
import math
import shutil
from pathlib import Path
from datetime import datetime

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE     = Path(r"d:\머신러닝 교육\ceria_pipeline_data")
XLSX     = BASE / "output" / "ceria_synthesis_database.xlsx"         # 원본 (파이프라인용, 절대 수정 금지)
XLSX_OUT = BASE / "output" / "ceria_synthesis_database_display.xlsx" # 서식 적용본 (Excel 열람용)
TEXT_DIR = BASE / "text"

SUMMARY_ROWS = 4               # 요약 영역 행 수 (1~4행)
COUNTS_ROW   = SUMMARY_ROWS + 2  # 열별 데이터 수 행 (6행)
DATA_HDR_ROW = SUMMARY_ROWS + 3  # 컬럼 헤더 행 (7행)

# ── 스타일 ────────────────────────────────────────────────────────────────────
_THIN  = Side(style="thin",   color="BFBFBF")
_MED   = Side(style="medium", color="4472C4")

BORDER_DATA   = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
BORDER_HEADER = Border(left=_MED,  right=_MED,  top=_MED,  bottom=_MED)

FILL_HEADER   = PatternFill("solid", fgColor="2E5FAC")   # 진파랑
FILL_SUMMARY  = PatternFill("solid", fgColor="DDEEFF")   # 연파랑
FILL_SUMTITLE = PatternFill("solid", fgColor="1F3864")   # 네이비
FILL_COUNTS   = PatternFill("solid", fgColor="FFF2CC")   # 연한 노랑 (데이터 수 행)
FILL_ODD      = PatternFill("solid", fgColor="FFFFFF")   # 흰색
FILL_EVEN     = PatternFill("solid", fgColor="F0F5FF")   # 아주 연한 파랑

FONT_TITLE    = Font(bold=True, size=15, color="FFFFFF")
FONT_LABEL    = Font(bold=True, size=10, color="2E5FAC")
FONT_VALUE    = Font(bold=True, size=11, color="1F3864")
FONT_UNIT     = Font(size=9,    color="666666")
FONT_COUNTS   = Font(bold=True, size=8,  color="7F6000")  # 데이터 수 행
FONT_HEADER   = Font(bold=True, size=9,  color="FFFFFF")
FONT_DATA     = Font(size=9,    color="000000")
FONT_LINK     = Font(size=9,    color="0563C1", underline="single")

ALIGN_CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=False)
ALIGN_LEFT    = Alignment(horizontal="left",   vertical="center", wrap_text=False)
ALIGN_LEFT_W  = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _load_xlsx(path) -> pd.DataFrame:
    """요약 행 유무·버전에 관계없이 'doi' 컬럼이 있는 행을 자동 감지하여 읽기."""
    raw = pd.read_excel(path, sheet_name=0, header=None, nrows=15)
    for idx, row in raw.iterrows():
        if any(str(v).strip().lower() == "doi" for v in row):
            df = pd.read_excel(path, sheet_name=0, header=idx)
            return df
    return pd.read_excel(path, sheet_name=0)   # 감지 실패 시 기본값

def _safe_fname(doi: str) -> str:
    return doi.replace("/", "_").replace(":", "_")

def has_text(doi: str) -> bool:
    return (TEXT_DIR / f"{_safe_fname(str(doi))}.txt").exists()

EXCEL_MAX_STR = 32_767  # Excel 셀 문자열 최대 길이

# 숫자 컬럼 → Excel 표시 포맷
NUM_FORMAT = {
    "year":                       "0",
    "citation_count":             "#,##0",
    "synthesis_temperature_c":    "0.0",
    "synthesis_time_h":           "0.0",
    "calcination_temperature_c":  "0.0",
    "calcination_time_h":         "0.0",
    "drying_temperature_c":       "0.0",
    "particle_size_tem_nm":       "0.00",
    "particle_size_sem_nm":       "0.00",
    "crystallite_size_xrd_nm":    "0.00",
    "bet_surface_area":           "0.0",
    "bet_surface_area_m2g":       "0.0",
    "ph_synthesis":               "0.0",
    "completeness_score":         "0.0",
    "dopant_concentration_mol_pct": "0.000",
}

def _to_excel(val):
    """pandas NA / NaN / None / numpy bool → Excel 호환 값으로 변환."""
    if val is None:
        return None
    # numpy / pandas bool → Python bool
    if isinstance(val, (bool, np.bool_)):
        return bool(val)
    # numpy 정수/실수 → Python 기본형
    if isinstance(val, np.integer):
        return int(val)
    if isinstance(val, np.floating):
        return None if np.isnan(val) else float(val)
    # pandas NA, NaT, NaN
    try:
        if pd.isna(val):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(val, float) and math.isnan(val):
        return None
    # 문자열 32,767자 초과 → 잘라냄 (Excel 셀 한도)
    if isinstance(val, str) and len(val) > EXCEL_MAX_STR:
        return val[:EXCEL_MAX_STR]
    return val


# ── 통계 계산 ─────────────────────────────────────────────────────────────────
print("통계 계산 중...")
df = _load_xlsx(XLSX)   # 요약 행 유무 자동 감지
total = len(df)

score_col = "completeness_score"
ml_ok = int(
    (pd.to_numeric(df.get(score_col, pd.Series(dtype=float)), errors="coerce")
     .fillna(0) >= 40).sum()
) if score_col in df.columns else 0

text_count = sum(1 for d in df["doi"].dropna() if has_text(str(d).strip()))

synth_count = int(
    df["is_synthesis_paper"].fillna(False).astype(bool).sum()
) if "is_synthesis_paper" in df.columns else 0

# composite 커버리지
size_cols = [c for c in ["particle_size_tem_nm", "crystallite_size_xrd_nm",
                          "particle_size_sem_nm"] if c in df.columns]
composite_pct = 0.0
if size_cols:
    has_size = df[size_cols].apply(pd.to_numeric, errors="coerce").notna().any(axis=1)
    composite_pct = has_size.mean() * 100

today = datetime.now().strftime("%Y-%m-%d %H:%M")

# 실제 데이터 기반 연도 범위 (하드코딩 대신 동적 계산)
if "year" in df.columns:
    year_series = pd.to_numeric(df["year"], errors="coerce").dropna()
    year_min = int(year_series.min()) if len(year_series) else 1990
    year_max = int(year_series.max()) if len(year_series) else datetime.now().year
else:
    year_min, year_max = 1990, datetime.now().year

# 연도 내림차순 정렬 (최신 논문 우선)
if "year" in df.columns:
    df = df.sort_values("year", ascending=False, na_position="last").reset_index(drop=True)

# 열별 비어있지 않은 데이터 수 (헤더 위 행에 표시용)
col_counts = {col: int(df[col].notna().sum()) for col in df.columns}

print(f"  총 논문:          {total:,}편")
print(f"  ML 가능 (≥40%):  {ml_ok:,}편")
print(f"  전문 보유:        {text_count:,}편")
print(f"  합성 논문:        {synth_count:,}편")
print(f"  크기 커버리지:    {composite_pct:.1f}%")


# ── 기존 시트 보존 (합성조건 외) ──────────────────────────────────────────────
print("\n기존 시트 로드...")
xl_orig = pd.ExcelFile(XLSX)
other_sheets = {}
for s in xl_orig.sheet_names:
    if s in ("합성조건", "Sheet1"):
        continue
    sdf = xl_orig.parse(s, header=None, nrows=15)
    # 각 보조 시트도 헤더 자동 감지
    hdr = 0
    for idx, row in sdf.iterrows():
        if any(str(v).strip().lower() not in ("nan", "") for v in row):
            hdr = idx
            break
    other_sheets[s] = xl_orig.parse(s, header=hdr)
xl_orig.close()


# ── 새 워크북 생성 ─────────────────────────────────────────────────────────────
print("서식 적용 중... (데이터가 많아 1~3분 소요)")
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "합성조건"

cols = list(df.columns)
n_cols = len(cols)

# ── 요약 블록 (행 1~4) ────────────────────────────────────────────────────────
TITLE_COL_END = min(n_cols, 8)

# 행1: 타이틀
ws.row_dimensions[1].height = 32
for c in range(1, n_cols + 1):
    ws.cell(1, c).fill = FILL_SUMTITLE
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TITLE_COL_END)
t = ws.cell(1, 1, f"CeO₂ 합성 논문 데이터베이스  ·  최종 업데이트: {today}")
t.font = FONT_TITLE
t.alignment = ALIGN_CENTER

# 행2: 첫번째 통계 줄
ws.row_dimensions[2].height = 24
stat1 = [
    ("총 논문",         f"{total:,} 편"),
    ("ML 가능 (≥40%)", f"{ml_ok:,} 편"),
    ("전문 보유",       f"{text_count:,} 편  ({text_count/total*100:.1f}%)"),
    ("합성 논문",       f"{synth_count:,} 편"),
]
col_idx = 1
for label, value in stat1:
    ws.cell(2, col_idx).fill     = FILL_SUMMARY
    ws.cell(2, col_idx).font     = FONT_LABEL
    ws.cell(2, col_idx).value    = label
    ws.cell(2, col_idx).alignment = ALIGN_CENTER
    col_idx += 1
    ws.cell(2, col_idx).fill     = FILL_SUMMARY
    ws.cell(2, col_idx).font     = FONT_VALUE
    ws.cell(2, col_idx).value    = value
    ws.cell(2, col_idx).alignment = ALIGN_CENTER
    col_idx += 1

# 나머지 열 채우기
for c in range(col_idx, n_cols + 1):
    ws.cell(2, c).fill = FILL_SUMMARY

# 행3: 두번째 통계 줄
ws.row_dimensions[3].height = 24
stat2 = [
    ("크기 커버리지",  f"{composite_pct:.1f} %"),
    ("비-OA (초록만)", f"{total - text_count:,} 편"),
    ("데이터 기간",    f"{year_min} ~ {year_max}"),
    ("출처",           "OpenAlex · Crossref · SemanticScholar"),
]
col_idx = 1
for label, value in stat2:
    ws.cell(3, col_idx).fill      = FILL_SUMMARY
    ws.cell(3, col_idx).font      = FONT_LABEL
    ws.cell(3, col_idx).value     = label
    ws.cell(3, col_idx).alignment = ALIGN_CENTER
    col_idx += 1
    ws.cell(3, col_idx).fill      = FILL_SUMMARY
    ws.cell(3, col_idx).font      = FONT_VALUE
    ws.cell(3, col_idx).value     = value
    ws.cell(3, col_idx).alignment = ALIGN_CENTER
    col_idx += 1

for c in range(col_idx, n_cols + 1):
    ws.cell(3, c).fill = FILL_SUMMARY

# 행4: 빈 줄 (구분선)
ws.row_dimensions[4].height = 6
for c in range(1, n_cols + 1):
    ws.cell(4, c).fill = PatternFill("solid", fgColor="2E5FAC")

# 행5: 빈 줄
ws.row_dimensions[5].height = 4

# ── 행6: 열별 데이터 수 (COUNTS_ROW) ─────────────────────────────────────────
ws.row_dimensions[COUNTS_ROW].height = 18
for ci, col_name in enumerate(cols, 1):
    cnt  = col_counts.get(col_name, 0)
    cell = ws.cell(COUNTS_ROW, ci)
    cell.value     = f"n={cnt:,}" if cnt > 0 else "-"
    cell.font      = FONT_COUNTS
    cell.fill      = FILL_COUNTS
    cell.alignment = ALIGN_CENTER
    cell.border    = BORDER_DATA

# ── 데이터 헤더 행 번호 (SUMMARY_ROWS 변경 시 자동 반영) ─────────────────────
HDR_ROW = DATA_HDR_ROW
ws.row_dimensions[HDR_ROW].height = 22
for ci, col_name in enumerate(cols, 1):
    cell = ws.cell(HDR_ROW, ci)
    cell.value     = col_name
    cell.font      = FONT_HEADER
    cell.fill      = FILL_HEADER
    cell.alignment = ALIGN_CENTER
    cell.border    = BORDER_HEADER

# ── 데이터 행 (행 7~) ─────────────────────────────────────────────────────────
print(f"  데이터 {total:,}행 작성 중...")
doi_ci     = cols.index("doi") + 1 if "doi" in cols else -1   # doi 컬럼 인덱스 (1-based)
fills      = [FILL_ODD, FILL_EVEN]                             # 홀/짝 교대 배경

for ri, (_, row) in enumerate(df.iterrows(), 1):
    excel_row = HDR_ROW + ri
    fill = fills[(ri - 1) % 2]
    ws.row_dimensions[excel_row].height = 15

    for ci, col_name in enumerate(cols, 1):
        val  = _to_excel(row[col_name])
        cell = ws.cell(excel_row, ci, val)        # value를 생성 시 동시에 설정
        cell.font      = FONT_LINK if ci == doi_ci else FONT_DATA
        cell.fill      = fill
        cell.alignment = ALIGN_LEFT
        cell.border    = BORDER_DATA
        # 숫자 컬럼 표시 포맷
        if col_name in NUM_FORMAT and val is not None:
            cell.number_format = NUM_FORMAT[col_name]

# ── 열 너비 자동 조정 (헤더 기준 + 샘플 데이터) ──────────────────────────────
print("  열 너비 조정 중...")
COL_MIN, COL_MAX = 8, 40
for ci, col_name in enumerate(cols, 1):
    # 헤더 길이
    max_len = len(str(col_name))
    # 처음 50행 샘플로 최대 길이 추정
    sample = df[col_name].dropna().astype(str).head(50)
    if len(sample):
        sample_max = sample.str.len().max()
        max_len = max(max_len, min(sample_max, COL_MAX))
    width = max(COL_MIN, min(max_len + 2, COL_MAX))
    ws.column_dimensions[get_column_letter(ci)].width = width

# ── 틀 고정 (헤더 바로 아래) ─────────────────────────────────────────────────
ws.freeze_panes = ws.cell(HDR_ROW + 1, 1)

# ── 자동 필터 ─────────────────────────────────────────────────────────────────
last_col_letter = get_column_letter(n_cols)
ws.auto_filter.ref = f"A{HDR_ROW}:{last_col_letter}{HDR_ROW}"

# ── 인쇄 설정 ─────────────────────────────────────────────────────────────────
ws.print_title_rows = f"{HDR_ROW}:{HDR_ROW}"
ws.page_setup.fitToWidth = 1

# ── 다른 시트 복원 ────────────────────────────────────────────────────────────
for sn, sdf in other_sheets.items():
    ws2 = wb.create_sheet(title=sn)
    for ci, col_name in enumerate(sdf.columns, 1):
        ws2.cell(1, ci).value = col_name
        ws2.cell(1, ci).font  = FONT_HEADER
        ws2.cell(1, ci).fill  = FILL_HEADER
    for ri, (_, row) in enumerate(sdf.iterrows(), 2):
        for ci, val in enumerate(row, 1):
            ws2.cell(ri, ci).value = _to_excel(val)

# ── 1. 서식 파일 저장 (열람용) ───────────────────────────────────────────────
print("\n저장 중...")
if XLSX_OUT.exists():
    shutil.copy2(XLSX_OUT, XLSX_OUT.with_suffix(".bak"))
wb.save(XLSX_OUT)
bak = XLSX_OUT.with_suffix(".bak")
if bak.exists():
    bak.unlink()
print(f"  서식 파일 저장 완료: {XLSX_OUT.name}")

# ── 2. 원본 복구 (파이프라인용 — 순수 데이터만, 요약 행 없음) ──────────────
print("  원본 Excel 복구 중 (요약 행 제거)...")
with pd.ExcelWriter(XLSX, engine="openpyxl") as w:
    df.to_excel(w, sheet_name="합성조건", index=False)
    for sn, sdf in other_sheets.items():
        sdf.to_excel(w, sheet_name=sn, index=False)
print(f"  원본 복구 완료: {XLSX.name}")

print(f"\n완료")
print(f"  열람용: {XLSX_OUT.name}  (요약+서식 적용, 연도 정렬)")
print(f"  원본:   {XLSX.name}      (순수 데이터, 파이프라인 전용)")
print(f"  데이터: {total:,}행  |  열: {len(cols)}개")


# ══════════════════════════════════════════════════════════════════════════════
# ceria_samples_merged.csv → ceria_samples_merged_display.xlsx
# ══════════════════════════════════════════════════════════════════════════════
SAMPLES_CSV  = BASE / "output" / "ceria_samples_merged.csv"
SAMPLES_OUT  = BASE / "output" / "ceria_samples_merged_display.xlsx"

# 초록 테마 (논문 DB 파란 테마와 구분)
FILL_S_TITLE  = PatternFill("solid", fgColor="1E5631")   # 진초록
FILL_S_SUMM   = PatternFill("solid", fgColor="D9EAD3")   # 연초록
FILL_S_HDR    = PatternFill("solid", fgColor="2D6A4F")   # 중간초록
FILL_S_COUNTS = PatternFill("solid", fgColor="EAF4EA")   # 아주 연한 초록
FILL_S_ODD    = PatternFill("solid", fgColor="FFFFFF")
FILL_S_EVEN   = PatternFill("solid", fgColor="F0FAF0")   # 아주 연한 초록
FONT_S_TITLE  = Font(bold=True, size=14, color="FFFFFF")
FONT_S_LABEL  = Font(bold=True, size=10, color="1E5631")
FONT_S_VALUE  = Font(bold=True, size=11, color="1E5631")
FONT_S_HDR    = Font(bold=True, size=9,  color="FFFFFF")
FONT_S_COUNTS = Font(bold=True, size=8,  color="1E5631")

if not SAMPLES_CSV.exists():
    print(f"\n[샘플 서식] {SAMPLES_CSV.name} 없음 — 건너뜀")
else:
    print(f"\n[샘플 서식] {SAMPLES_CSV.name} 로드 중...")
    ds = pd.read_csv(SAMPLES_CSV, dtype=str, low_memory=False)
    s_total   = len(ds)
    s_papers  = ds["doi"].nunique() if "doi" in ds.columns else 0

    def _s_cov(col):
        if col not in ds.columns:
            return 0, 0.0
        n = ds[col].replace("", pd.NA).notna().sum()
        return int(n), n / s_total * 100 if s_total else 0.0

    n_tem,  p_tem  = _s_cov("particle_size_tem_nm")
    n_sem,  p_sem  = _s_cov("particle_size_sem_nm")
    n_xrd,  p_xrd  = _s_cov("crystallite_size_xrd_nm")
    n_meth, p_meth = _s_cov("synthesis_method")
    n_prec, p_prec = _s_cov("ce_precursor")
    n_solv, p_solv = _s_cov("solvent")

    # TEM or SEM 통합 1차입자
    tem_s = ds["particle_size_tem_nm"].replace("", pd.NA) if "particle_size_tem_nm" in ds.columns else pd.Series(pd.NA, index=ds.index)
    sem_s = ds["particle_size_sem_nm"].replace("", pd.NA) if "particle_size_sem_nm" in ds.columns else pd.Series(pd.NA, index=ds.index)
    n_primary = int((tem_s.notna() | sem_s.notna()).sum())
    p_primary = n_primary / s_total * 100 if s_total else 0.0

    # 숫자 컬럼 변환
    for nc in ["particle_size_tem_nm", "particle_size_sem_nm",
               "crystallite_size_xrd_nm", "bet_surface_area_m2g",
               "synthesis_temperature_c", "synthesis_time_h",
               "calcination_temperature_c", "calcination_time_h",
               "ph_synthesis", "dopant_concentration_mol_pct"]:
        if nc in ds.columns:
            ds[nc] = pd.to_numeric(ds[nc], errors="coerce")

    # doi 기준 정렬 (논문 묶음으로 보기 편하게)
    if "doi" in ds.columns:
        ds = ds.sort_values("doi", na_position="last").reset_index(drop=True)

    s_cols   = list(ds.columns)
    s_ncols  = len(s_cols)
    s_col_counts = {col: int(ds[col].notna().sum()) for col in s_cols}

    print(f"  {s_total:,}개 샘플 / {s_papers:,}편 논문")

    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "샘플추출결과"

    S_SUMM_ROWS  = 3
    S_COUNTS_ROW = S_SUMM_ROWS + 2
    S_HDR_ROW    = S_SUMM_ROWS + 3
    TC = min(s_ncols, 8)

    # 행1: 타이틀
    ws2.row_dimensions[1].height = 30
    for c in range(1, s_ncols + 1):
        ws2.cell(1, c).fill = FILL_S_TITLE
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TC)
    t = ws2.cell(1, 1, f"CeO₂ 샘플 추출 결과  ·  최종 업데이트: {today}")
    t.font = FONT_S_TITLE; t.alignment = ALIGN_CENTER

    # 행2: 통계 (샘플수 / 논문수 / 합성법 / Ce전구체 / 용매)
    ws2.row_dimensions[2].height = 22
    stat_s1 = [
        ("총 샘플",   f"{s_total:,} 개"),
        ("대상 논문", f"{s_papers:,} 편"),
        ("합성방법",  f"{n_meth:,}  ({p_meth:.1f}%)"),
        ("Ce 전구체", f"{n_prec:,}  ({p_prec:.1f}%)"),
    ]
    ci = 1
    for lbl, val in stat_s1:
        ws2.cell(2, ci).fill = FILL_S_SUMM; ws2.cell(2, ci).font = FONT_S_LABEL
        ws2.cell(2, ci).value = lbl; ws2.cell(2, ci).alignment = ALIGN_CENTER; ci += 1
        ws2.cell(2, ci).fill = FILL_S_SUMM; ws2.cell(2, ci).font = FONT_S_VALUE
        ws2.cell(2, ci).value = val; ws2.cell(2, ci).alignment = ALIGN_CENTER; ci += 1
    for c in range(ci, s_ncols + 1): ws2.cell(2, c).fill = FILL_S_SUMM

    # 행3: 입자크기 커버리지
    ws2.row_dimensions[3].height = 22
    stat_s2 = [
        ("1차입자 (TEM|SEM)", f"{n_primary:,}  ({p_primary:.1f}%)"),
        ("XRD 결정립",        f"{n_xrd:,}  ({p_xrd:.1f}%)"),
        ("용매",              f"{n_solv:,}  ({p_solv:.1f}%)"),
        ("TEM",               f"{n_tem:,}  ({p_tem:.1f}%)"),
    ]
    ci = 1
    for lbl, val in stat_s2:
        ws2.cell(3, ci).fill = FILL_S_SUMM; ws2.cell(3, ci).font = FONT_S_LABEL
        ws2.cell(3, ci).value = lbl; ws2.cell(3, ci).alignment = ALIGN_CENTER; ci += 1
        ws2.cell(3, ci).fill = FILL_S_SUMM; ws2.cell(3, ci).font = FONT_S_VALUE
        ws2.cell(3, ci).value = val; ws2.cell(3, ci).alignment = ALIGN_CENTER; ci += 1
    for c in range(ci, s_ncols + 1): ws2.cell(3, c).fill = FILL_S_SUMM

    # 행4: 구분선
    ws2.row_dimensions[4].height = 5
    for c in range(1, s_ncols + 1):
        ws2.cell(4, c).fill = FILL_S_HDR

    # 행5: 빈줄
    ws2.row_dimensions[5].height = 4

    # 행6: 열별 데이터 수
    ws2.row_dimensions[S_COUNTS_ROW].height = 16
    for ci, col in enumerate(s_cols, 1):
        cnt = s_col_counts.get(col, 0)
        cell = ws2.cell(S_COUNTS_ROW, ci)
        cell.value = f"n={cnt:,}" if cnt > 0 else "-"
        cell.font = FONT_S_COUNTS; cell.fill = FILL_S_COUNTS
        cell.alignment = ALIGN_CENTER; cell.border = BORDER_DATA

    # 행7: 헤더
    ws2.row_dimensions[S_HDR_ROW].height = 20
    for ci, col in enumerate(s_cols, 1):
        cell = ws2.cell(S_HDR_ROW, ci)
        cell.value = col; cell.font = FONT_S_HDR
        cell.fill = FILL_S_HDR; cell.alignment = ALIGN_CENTER; cell.border = BORDER_HEADER

    # 데이터 행
    print(f"  데이터 {s_total:,}행 작성 중...")
    doi_ci2 = s_cols.index("doi") + 1 if "doi" in s_cols else -1
    fills2  = [FILL_S_ODD, FILL_S_EVEN]
    for ri, (_, row) in enumerate(ds.iterrows(), 1):
        excel_row = S_HDR_ROW + ri
        fill = fills2[(ri - 1) % 2]
        ws2.row_dimensions[excel_row].height = 15
        for ci, col in enumerate(s_cols, 1):
            val  = _to_excel(row[col])
            cell = ws2.cell(excel_row, ci, val)
            cell.font      = FONT_LINK if ci == doi_ci2 else FONT_DATA
            cell.fill      = fill
            cell.alignment = ALIGN_LEFT
            cell.border    = BORDER_DATA
            if col in NUM_FORMAT and val is not None:
                cell.number_format = NUM_FORMAT[col]

    # 열 너비
    for ci, col in enumerate(s_cols, 1):
        max_len = len(str(col))
        sample  = ds[col].dropna().astype(str).head(50)
        if len(sample):
            max_len = max(max_len, min(sample.str.len().max(), 40))
        ws2.column_dimensions[get_column_letter(ci)].width = max(8, min(max_len + 2, 40))

    # 틀 고정 + 자동 필터
    ws2.freeze_panes = ws2.cell(S_HDR_ROW + 1, 1)
    ws2.auto_filter.ref = f"A{S_HDR_ROW}:{get_column_letter(s_ncols)}{S_HDR_ROW}"

    # 저장
    if SAMPLES_OUT.exists():
        SAMPLES_OUT.unlink()
    wb2.save(SAMPLES_OUT)
    print(f"  저장 완료: {SAMPLES_OUT.name}  ({s_total:,}샘플 × {s_ncols}열)")
