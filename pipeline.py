# %% [셀 1] 환경 점검 - 패키지 및 API 키 확인
"""
실행 순서: 이 셀부터 순서대로 실행하세요.
에러 발생 시 해당 셀에서 멈추고 메시지를 확인하세요.
"""
import sys
import os
import importlib

print(f"Python: {sys.version}")
print(f"작업 디렉토리: {os.getcwd()}")

try:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
except NameError:
    BASE_DIR = os.getcwd()

print(f"BASE_DIR: {BASE_DIR}")

REQUIRED_PACKAGES = [
    ("requests",   "requests"),
    ("pandas",     "pandas"),
    ("openpyxl",   "openpyxl"),
    ("pdfplumber", "pdfplumber"),
    ("bs4",        "beautifulsoup4"),
    ("tqdm",       "tqdm"),
    ("dotenv",     "python-dotenv"),
]
OPTIONAL_PACKAGES = [
    ("anthropic",  "anthropic"),
    ("fitz",       "PyMuPDF"),
]

missing = []
for import_name, install_name in REQUIRED_PACKAGES:
    try:
        importlib.import_module(import_name)
        print(f"  [OK] {import_name}")
    except ImportError:
        print(f"  [없음] {install_name} → 필수")
        missing.append(install_name)

for import_name, install_name in OPTIONAL_PACKAGES:
    try:
        importlib.import_module(import_name)
        print(f"  [OK] {import_name} (선택)")
    except ImportError:
        print(f"  [없음] {install_name} → 선택사항")

if missing:
    print(f"\n필수 패키지 설치:\n  pip install {' '.join(missing)}")
else:
    print("\n필수 패키지 모두 OK")


# %% [셀 2] .env 로드 및 API 키 확인
from dotenv import load_dotenv

ENV_PATH = os.path.join(BASE_DIR, ".env")
if os.path.exists(ENV_PATH):
    load_dotenv(ENV_PATH)
    print(f".env 로드됨: {ENV_PATH}")
else:
    print(".env 파일 없음 → 아래 직접 입력 가능")
    # os.environ["ANTHROPIC_API_KEY"] = "sk-ant-..."
    # os.environ["UNPAYWALL_EMAIL"]   = "juhur@soulbrain.co.kr"

ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
OPENAI_API_KEY    = os.environ.get("OPENAI_API_KEY", "")
UNPAYWALL_EMAIL   = os.environ.get("UNPAYWALL_EMAIL", "juhur@soulbrain.co.kr")
YOUR_EMAIL        = os.environ.get("YOUR_EMAIL", UNPAYWALL_EMAIL)

print(f"ANTHROPIC_API_KEY : {'설정됨' if ANTHROPIC_API_KEY else '없음'}")
print(f"OPENAI_API_KEY    : {'설정됨' if OPENAI_API_KEY else '없음 (OpenAI 추출 건너뜀)'}")
print(f"UNPAYWALL_EMAIL   : {UNPAYWALL_EMAIL}")


# %% [셀 3] 검색 설정
import time
import re
import json

# ===== 검색 파라미터 =====
SEARCH_QUERIES = [
    # ── 기본 합성 쿼리 ──────────────────────────────────────────────────────────
    "ceria nanoparticles synthesis",
    "CeO2 nanoparticles synthesis",
    "cerium oxide nanoparticles hydrothermal",
    "cerium oxide nanoparticles sol-gel",
    "ceria nanoparticles precipitation",
    "CeO2 nanoparticles solvothermal",
    "cerium oxide nanoparticles combustion",
    "ceria nanoparticles morphology",
    # ── 합성법 다양화 ───────────────────────────────────────────────────────────
    "cerium oxide nanoparticles microwave synthesis",
    "cerium oxide nanoparticles template synthesis",
    "ceria nanoparticles thermal decomposition synthesis",
    "cerium oxide nanoparticles spray pyrolysis",
    "ceria nanoparticles sonochemical synthesis",
    # ── 도핑/기능화 ─────────────────────────────────────────────────────────────
    "doped ceria nanoparticles synthesis",
    "CeO2 nanoparticles rare earth doped",
    "zirconia ceria nanoparticles synthesis",
    # ── 응용/특성 기반 (합성 정보 포함) ──────────────────────────────────────────
    "CeO2 nanoparticles photocatalysis synthesis",
    "ceria nanoparticles catalysis preparation",
    "cerium oxide nanoparticles green synthesis",
    "CeO2 nanoparticles biomedical synthesis",
    # ── 특수 형태 ───────────────────────────────────────────────────────────────
    "cerium oxide quantum dots synthesis",
    "ceria nanorods nanocubes synthesis",
    "porous ceria nanoparticles synthesis",
]

YEAR_START = 1990          # 수집 시작 연도
YEAR_END   = 2026          # 수집 종료 연도

# 소스별 최대 수집
MAX_PER_QUERY_OPENALEX       = 500   # OpenAlex: 쿼리당 최대
MAX_PER_QUERY_CROSSREF       = 200   # Crossref: 쿼리당 최대
MAX_PER_QUERY_SEMANTIC       = 200   # Semantic Scholar: 쿼리당 최대
MAX_TOTAL_PAPERS             = 10000 # 전체 최대 (DOI 중복 제거 전, 8→24개 쿼리 반영)

# Sci-Hub 사용 여부 (True = 비공개 논문 PDF 자동 수집 시도)
SCIHUB_ENABLED = False
SCIHUB_URLS = [
    "https://sci-hub.mk/",
    "https://sci-hub.se/",
    "https://sci-hub.st/",
]

OUTPUT_DIR = os.path.join(BASE_DIR, "output")
PDF_DIR    = os.path.join(BASE_DIR, "pdf")
TEXT_DIR   = os.path.join(BASE_DIR, "text")
for d in [OUTPUT_DIR, PDF_DIR, TEXT_DIR]:
    os.makedirs(d, exist_ok=True)

print(f"수집 연도: {YEAR_START} ~ {YEAR_END}")
print(f"검색어: {len(SEARCH_QUERIES)}개")
print(f"전체 최대: {MAX_TOTAL_PAPERS}편")
print(f"Sci-Hub: {'활성화' if SCIHUB_ENABLED else '비활성화'}")


# %% [셀 4] OpenAlex 수집 (cursor 기반 페이지네이션)
"""
OpenAlex는 무료·무키·고품질 API로 전체 논문의 ~90% 커버.
cursor 방식으로 10,000편까지 수집 가능.
초록은 inverted_index 형식 → 자동으로 평문 복원.
"""
import requests
from tqdm import tqdm

OPENALEX_BASE = "https://api.openalex.org/works"
OPENALEX_FIELDS = (
    "id,doi,title,authorships,publication_year,publication_date,"
    "primary_location,best_oa_location,open_access,abstract_inverted_index,"
    "cited_by_count,type"
)

METHOD_KEYWORDS = {
    "hydrothermal":        ["hydrothermal"],
    "solvothermal":        ["solvothermal"],
    "precipitation":       ["precipitation", "co-precipitation", "coprecipitation"],
    "sol-gel":             ["sol-gel", "sol gel"],
    "thermal_decomp":      ["thermal decomposition", "thermolysis"],
    "microemulsion":       ["microemulsion", "reverse micelle"],
    "combustion":          ["combustion synthesis", "solution combustion"],
    "polyol":              ["polyol process", "ethylene glycol", "butanediol"],
    "sonochemical":        ["sonochem", "ultrasonic", "ultrasound", "sonication"],
    "microwave":           ["microwave"],
    "spray_pyrolysis":     ["spray pyrolysis"],
    "template":            ["template", "hard template", "soft template"],
    "green":               ["green synthesis", "plant extract", "biogenic"],
}

MORPHOLOGY_KEYWORDS = {
    "nanoparticle":  ["nanoparticle", "nanocrystal", "quantum dot"],
    "nanorod":       ["nanorod", "nanowire", "nanotube"],
    "nanocube":      ["nanocube", "cube", "cubic"],
    "nanosphere":    ["nanosphere", "spherical"],
    "nanoflower":    ["nanoflower", "hierarchical", "flower-like"],
    "octahedra":     ["octahedr", "polyhedr"],
    "porous":        ["porous", "mesoporous", "hollow"],
    "nanosheet":     ["nanosheet", "nanoplate", "2d"],
}

# ── §5 확장 triage 태그 (stage1 가이드) ───────────────────────────────────────
MINERALIZER_KEYWORDS = {
    "NaOH":   ["naoh", "sodium hydroxide"],
    "KOH":    ["koh", "potassium hydroxide"],
    "NH3":    ["ammonia", "ammonium hydroxide", "nh4oh", "nh3"],
    "urea":   ["urea"],
    "HMTA":   ["hexamethylenetetramine", "hmta", "hexamine"],
    "TMAH":   ["tmah", "tetramethylammonium hydroxide"],
    "Na2CO3": ["sodium carbonate", "na2co3"],
}

ADDITIVE_KEYWORDS = {
    "CTAB":       ["ctab", "cetyltrimethylammonium"],
    "SDS":        ["sds", "sodium dodecyl sulfate", "dodecyl sulphate"],
    "PVP":        ["pvp", "polyvinylpyrrolidone", "polyvinyl pyrrolidone"],
    "PEG":        ["peg", "polyethylene glycol"],
    "triton":     ["triton x-100", "triton"],
    "pluronic":   ["pluronic", "p123", "f127"],
    "citrate":    ["citrate", "citric acid"],
    "oleic_acid": ["oleic acid", "oleate"],
    "oleylamine": ["oleylamine"],
    "EDTA":       ["edta", "ethylenediaminetetraacetic"],
    "DEA":        ["diethanolamine"],
    "TEA":        ["triethanolamine"],
}

SOLVENT_KEYWORDS = {
    "water":          ["aqueous", "deionized water", "distilled water"],
    "ethylene_glycol":["ethylene glycol", "polyol"],
    "ethanol":        ["ethanol"],
    "methanol":       ["methanol"],
    "butanediol":     ["butanediol"],
    "isopropanol":    ["isopropanol", "2-propanol", " ipa "],
}

OXIDANT_ASSIST_KEYWORDS = {
    "H2O2":        ["h2o2", "hydrogen peroxide"],
    "sonochemical":["sonochem", "ultrasonic", "ultrasound", "sonication"],
    "microwave":   ["microwave"],
    "calcination": ["calcin", "anneal"],
}

DOPANT_TAG_KEYWORDS = {
    "La": ["la-doped", "lanthanum doped", "la doped"],
    "Nd": ["nd-doped", "neodymium doped"],
    "Pr": ["pr-doped", "praseodymium doped"],
    "Sm": ["sm-doped", "samarium doped"],
    "Y":  ["y-doped", "yttrium doped"],
    "Gd": ["gd-doped", "gadolinium doped"],
    "Eu": ["eu-doped", "europium doped"],
    "Zr": ["zr-doped", "zirconia doped", "zirconium doped"],
    "Co": ["co-doped ceria", "cobalt doped"],
    "Fe": ["fe-doped", "iron doped"],
}

def _tag_keywords(text: str, keyword_map: dict) -> str:
    low = text.lower()
    matched = [label for label, variants in keyword_map.items()
               if any(v in low for v in variants)]
    return "|".join(matched)


def _reconstruct_abstract(inverted_index: dict) -> str:
    if not inverted_index:
        return ""
    words = {}
    for word, positions in inverted_index.items():
        for pos in positions:
            words[pos] = word
    return " ".join(words[i] for i in sorted(words))

def _openalex_paper_to_row(work: dict) -> dict:
    doi_raw = work.get("doi", "") or ""
    doi = doi_raw.replace("https://doi.org/", "").strip().lower()

    authors = work.get("authorships", []) or []
    author_names = [
        (a.get("author") or {}).get("display_name", "")
        for a in authors[:6]
    ]
    author_str = "; ".join(n for n in author_names if n)
    if len(authors) > 6:
        author_str += f" et al. (+{len(authors)-6})"

    loc = work.get("primary_location") or {}
    source = loc.get("source") or {}
    journal = source.get("display_name", "")
    pdf_url = loc.get("pdf_url", "") or ""

    oa = work.get("open_access") or {}
    best_oa = work.get("best_oa_location") or {}
    if not pdf_url:
        pdf_url = best_oa.get("pdf_url", "") or oa.get("oa_url", "") or ""

    abstract = _reconstruct_abstract(work.get("abstract_inverted_index") or {})
    title = work.get("title", "") or ""
    haystack = f"{title} {abstract}"

    return {
        "source_api":         "OpenAlex",
        "source_id":          work.get("id", ""),
        "doi":                doi,
        "title":              title,
        "authors":            author_str,
        "year":               work.get("publication_year"),
        "pub_date":           work.get("publication_date", ""),
        "journal":            journal,
        "citation_count":     work.get("cited_by_count", 0),
        "abstract":           abstract,
        "open_access_url":    pdf_url,
        "is_oa":               oa.get("is_oa", False),
        "oa_status":           oa.get("oa_status", ""),
        "tagged_methods":      _tag_keywords(haystack, METHOD_KEYWORDS),
        "tagged_morphologies": _tag_keywords(haystack, MORPHOLOGY_KEYWORDS),
        "tagged_mineralizer":  _tag_keywords(haystack, MINERALIZER_KEYWORDS),
        "tagged_additives":    _tag_keywords(haystack, ADDITIVE_KEYWORDS),
        "tagged_solvent":      _tag_keywords(haystack, SOLVENT_KEYWORDS),
        "tagged_assist":       _tag_keywords(haystack, OXIDANT_ASSIST_KEYWORDS),
        "tagged_dopant":       _tag_keywords(haystack, DOPANT_TAG_KEYWORDS),
    }


def collect_openalex(query: str, year_start: int, year_end: int,
                     max_per_query: int = 500) -> list:
    collected = []
    cursor = "*"
    per_page = 200
    filters = f"publication_year:{year_start}-{year_end}"

    while len(collected) < max_per_query:
        params = {
            "search":   query,
            "filter":   filters,
            "per-page": per_page,
            "cursor":   cursor,
            "select":   OPENALEX_FIELDS,
        }
        if YOUR_EMAIL:
            params["mailto"] = YOUR_EMAIL

        data = None
        for attempt in range(4):
            try:
                resp = requests.get(OPENALEX_BASE, params=params,
                                    timeout=30,
                                    headers={"User-Agent": f"ceria-pipeline/1.0 (mailto:{YOUR_EMAIL})"})
                if resp.status_code in (429, 500, 502, 503):
                    time.sleep(2 ** attempt)
                    continue
                resp.raise_for_status()
                data = resp.json()
                break
            except Exception as e:
                print(f"    OpenAlex 오류(시도{attempt+1}): {e}")
                time.sleep(2 ** attempt)
        if data is None:
            break

        results = data.get("results", [])
        if not results:
            break

        for work in results:
            collected.append(_openalex_paper_to_row(work))

        meta = data.get("meta", {})
        cursor = meta.get("next_cursor")
        if not cursor:
            break

        time.sleep(0.15)

    return collected


openalex_papers = []
print("=== OpenAlex 수집 ===")
for q in SEARCH_QUERIES:
    print(f"  쿼리: '{q}'")
    batch = collect_openalex(q, YEAR_START, YEAR_END, MAX_PER_QUERY_OPENALEX)
    openalex_papers.extend(batch)
    print(f"    → {len(batch)}편 (누적: {len(openalex_papers)}편)")
    if len(openalex_papers) >= MAX_TOTAL_PAPERS:
        break
    time.sleep(0.5)

print(f"\nOpenAlex 총 수집: {len(openalex_papers)}편 (DOI 중복 포함)")


# %% [셀 5] Crossref 수집
"""
Crossref는 DOI 레지스트리로 메타데이터 품질이 높음.
abstract가 없는 경우가 많지만 정확한 DOI·출판일 확보에 유용.
"""

CROSSREF_BASE = "https://api.crossref.org/works"

def _crossref_paper_to_row(item: dict) -> dict:
    doi = (item.get("DOI", "") or "").strip().lower()

    title_list = item.get("title", []) or []
    title = title_list[0] if title_list else ""

    authors = item.get("author", []) or []
    name_parts = []
    for a in authors[:6]:
        given = a.get("given", "")
        family = a.get("family", "")
        name_parts.append(f"{given} {family}".strip())
    author_str = "; ".join(name_parts)
    if len(authors) > 6:
        author_str += f" et al. (+{len(authors)-6})"

    issued = item.get("issued", {})
    date_parts = issued.get("date-parts", [[]])
    year = date_parts[0][0] if date_parts and date_parts[0] else None
    month = date_parts[0][1] if date_parts and len(date_parts[0]) > 1 else ""
    pub_date = f"{year}-{month:02d}" if year and month else str(year or "")

    container = item.get("container-title", []) or []
    journal = container[0] if container else ""

    abstract_raw = item.get("abstract", "") or ""
    abstract = re.sub(r"<[^>]+>", " ", abstract_raw).strip()

    links = item.get("link", []) or []
    pdf_url = ""
    for link in links:
        if "pdf" in (link.get("content-type", "") or "").lower():
            pdf_url = link.get("URL", "")
            break

    haystack = f"{title} {abstract}"

    return {
        "source_api":          "Crossref",
        "source_id":           doi,
        "doi":                 doi,
        "title":               title,
        "authors":             author_str,
        "year":                year,
        "pub_date":            pub_date,
        "journal":             journal,
        "citation_count":      item.get("is-referenced-by-count", 0),
        "abstract":            abstract,
        "open_access_url":     pdf_url,
        "is_oa":               bool(pdf_url),
        "oa_status":           "oa" if pdf_url else "",
        "tagged_methods":      _tag_keywords(haystack, METHOD_KEYWORDS),
        "tagged_morphologies": _tag_keywords(haystack, MORPHOLOGY_KEYWORDS),
        "tagged_mineralizer":  _tag_keywords(haystack, MINERALIZER_KEYWORDS),
        "tagged_additives":    _tag_keywords(haystack, ADDITIVE_KEYWORDS),
        "tagged_solvent":      _tag_keywords(haystack, SOLVENT_KEYWORDS),
        "tagged_assist":       _tag_keywords(haystack, OXIDANT_ASSIST_KEYWORDS),
        "tagged_dopant":       _tag_keywords(haystack, DOPANT_TAG_KEYWORDS),
    }


def collect_crossref(query: str, year_start: int, year_end: int,
                     max_per_query: int = 200) -> list:
    collected = []
    rows_per_page = 100
    offset = 0

    while len(collected) < max_per_query:
        params = {
            "query":  query,
            "filter": f"from-pub-date:{year_start},until-pub-date:{year_end}",
            "rows":   rows_per_page,
            "offset": offset,
            "select": "DOI,title,author,issued,container-title,abstract,is-referenced-by-count,link",
        }
        if YOUR_EMAIL:
            params["mailto"] = YOUR_EMAIL

        try:
            resp = requests.get(CROSSREF_BASE, params=params, timeout=30)
            resp.raise_for_status()
            data = resp.json()
        except Exception as e:
            print(f"    Crossref 오류: {e}")
            break

        items = data.get("message", {}).get("items", []) or []
        if not items:
            break

        for item in items:
            collected.append(_crossref_paper_to_row(item))

        offset += len(items)
        if len(items) < rows_per_page:
            break
        time.sleep(0.2)

    return collected


crossref_papers = []
print("=== Crossref 수집 ===")
for q in SEARCH_QUERIES:
    print(f"  쿼리: '{q}'")
    batch = collect_crossref(q, YEAR_START, YEAR_END, MAX_PER_QUERY_CROSSREF)
    crossref_papers.extend(batch)
    print(f"    → {len(batch)}편 (누적: {len(crossref_papers)}편)")
    if len(crossref_papers) >= MAX_TOTAL_PAPERS:
        break
    time.sleep(0.5)

print(f"\nCrossref 총 수집: {len(crossref_papers)}편 (DOI 중복 포함)")


# %% [셀 6] Semantic Scholar 수집
SEMANTIC_SCHOLAR_BASE = "https://api.semanticscholar.org/graph/v1"
SS_FIELDS = "paperId,title,authors,year,externalIds,openAccessPdf,abstract,venue,citationCount,publicationDate"

def _ss_paper_to_row(paper: dict) -> dict:
    ext = paper.get("externalIds") or {}
    doi = (ext.get("DOI", "") or "").strip().lower()

    authors = paper.get("authors", []) or []
    author_str = "; ".join(a.get("name", "") for a in authors[:6])
    if len(authors) > 6:
        author_str += f" et al. (+{len(authors)-6})"

    oa = paper.get("openAccessPdf") or {}
    pdf_url = oa.get("url", "") or ""

    title = paper.get("title", "") or ""
    abstract_ss = (paper.get("abstract") or "")
    haystack = f"{title} {abstract_ss}"

    return {
        "source_api":          "SemanticScholar",
        "source_id":           paper.get("paperId", ""),
        "doi":                 doi,
        "title":               title,
        "authors":             author_str,
        "year":                paper.get("year"),
        "pub_date":            paper.get("publicationDate", ""),
        "journal":             paper.get("venue", ""),
        "citation_count":      paper.get("citationCount", 0),
        "abstract":            abstract_ss,
        "open_access_url":     pdf_url,
        "is_oa":               bool(pdf_url),
        "oa_status":           "oa" if pdf_url else "",
        "tagged_methods":      _tag_keywords(haystack, METHOD_KEYWORDS),
        "tagged_morphologies": _tag_keywords(haystack, MORPHOLOGY_KEYWORDS),
        "tagged_mineralizer":  _tag_keywords(haystack, MINERALIZER_KEYWORDS),
        "tagged_additives":    _tag_keywords(haystack, ADDITIVE_KEYWORDS),
        "tagged_solvent":      _tag_keywords(haystack, SOLVENT_KEYWORDS),
        "tagged_assist":       _tag_keywords(haystack, OXIDANT_ASSIST_KEYWORDS),
        "tagged_dopant":       _tag_keywords(haystack, DOPANT_TAG_KEYWORDS),
    }


def collect_semantic_scholar(query: str, year_start: int, year_end: int,
                              max_per_query: int = 200) -> list:
    collected = []
    offset = 0

    while len(collected) < max_per_query:
        params = {
            "query":  query,
            "fields": SS_FIELDS,
            "limit":  100,
            "offset": offset,
            "year":   f"{year_start}-{year_end}",
        }
        try:
            resp = requests.get(
                f"{SEMANTIC_SCHOLAR_BASE}/paper/search",
                params=params, timeout=30
            )
            resp.raise_for_status()
            data = resp.json()
        except Exception as e:
            print(f"    SemanticScholar 오류: {e}")
            break

        batch = data.get("data", []) or []
        if not batch:
            break

        collected.extend(_ss_paper_to_row(p) for p in batch)
        offset += len(batch)
        if len(batch) < 100:
            break
        time.sleep(1.0)

    return collected


semantic_papers = []
print("=== Semantic Scholar 수집 ===")
for q in SEARCH_QUERIES:
    print(f"  쿼리: '{q}'")
    batch = collect_semantic_scholar(q, YEAR_START, YEAR_END, MAX_PER_QUERY_SEMANTIC)
    semantic_papers.extend(batch)
    print(f"    → {len(batch)}편 (누적: {len(semantic_papers)}편)")
    if len(semantic_papers) >= MAX_TOTAL_PAPERS:
        break
    time.sleep(1.0)

print(f"\nSemantic Scholar 총 수집: {len(semantic_papers)}편 (DOI 중복 포함)")


# %% [셀 7] DOI 기반 통합 중복 제거
"""
1순위: DOI 동일 → 하나만 남김 (OpenAlex > Crossref > SemanticScholar 순 우선)
2순위: DOI 없음 → 제목 유사도로 중복 제거 (소문자 정규화 후 exact match)
"""
import pandas as pd

all_raw = openalex_papers + crossref_papers + semantic_papers
print(f"전체 수집: {len(all_raw)}편")

# source 우선순위
SOURCE_PRIORITY = {"OpenAlex": 0, "Crossref": 1, "SemanticScholar": 2}

def _norm_doi(doi: str) -> str:
    return (doi or "").strip().lower().replace("https://doi.org/", "")

def _norm_title(title: str) -> str:
    if not title:
        return ""
    return re.sub(r"\s+", " ", title.lower().strip())

# DOI 기반 중복 제거
doi_map: dict = {}       # doi → row (best source)
no_doi_list: list = []

for row in all_raw:
    doi = _norm_doi(row.get("doi", ""))
    row["doi"] = doi
    if not doi:
        no_doi_list.append(row)
        continue
    if doi not in doi_map:
        doi_map[doi] = row
    else:
        existing_pri = SOURCE_PRIORITY.get(doi_map[doi]["source_api"], 99)
        new_pri      = SOURCE_PRIORITY.get(row["source_api"], 99)
        if new_pri < existing_pri:
            # 더 좋은 소스로 교체하되, 기존 항목의 abstract/url이 있으면 보완
            if not row.get("abstract") and doi_map[doi].get("abstract"):
                row["abstract"] = doi_map[doi]["abstract"]
            if not row.get("open_access_url") and doi_map[doi].get("open_access_url"):
                row["open_access_url"] = doi_map[doi]["open_access_url"]
            doi_map[doi] = row

# DOI 없는 논문 제목 기반 중복 제거
title_map: dict = {}
for row in no_doi_list:
    nt = _norm_title(row.get("title", ""))
    if nt and nt not in title_map:
        title_map[nt] = row

# 전체 unique papers
unique_papers = list(doi_map.values()) + list(title_map.values())

# DOI 있는 논문끼리 abstract/url 보완 (OpenAlex는 abstract 있고 Crossref는 URL 있는 경우)
# → 이미 위에서 best-source를 저장했으므로 Crossref의 URL 보완만 추가
for p in crossref_papers:
    doi = _norm_doi(p.get("doi", ""))
    if doi and doi in doi_map:
        existing = doi_map[doi]
        if not existing.get("open_access_url") and p.get("open_access_url"):
            existing["open_access_url"] = p["open_access_url"]
        if not existing.get("abstract") and p.get("abstract"):
            existing["abstract"] = p["abstract"]

print(f"중복 제거 후: {len(unique_papers)}편")
print(f"  DOI 있음: {len(doi_map)}편")
print(f"  DOI 없음(제목 기반): {len(title_map)}편")

# pandas 변환 및 중간 저장
df_meta = pd.DataFrame(unique_papers).reset_index(drop=True)
df_meta["paper_id"] = df_meta.apply(
    lambda r: r["doi"] if r["doi"] else f"NOID_{r.name:05d}", axis=1
)

meta_path = os.path.join(OUTPUT_DIR, "papers_metadata.xlsx")
df_meta.to_excel(meta_path, index=False)
print(f"메타데이터 저장: {meta_path}")
print(df_meta[["title", "year", "doi", "source_api", "citation_count"]].head(10).to_string())


# %% [셀 8] Unpaywall OA PDF URL 보완
"""
DOI가 있는 논문에 대해 Unpaywall API로 합법적인 OA PDF URL을 조회합니다.
"""
UNPAYWALL_BASE = "https://api.unpaywall.org/v2"

def get_unpaywall_pdf(doi: str, email: str) -> str:
    if not doi or not email:
        return ""
    url = f"{UNPAYWALL_BASE}/{requests.utils.quote(doi, safe='')}?email={email}"
    try:
        resp = requests.get(url, timeout=15)
        if resp.status_code == 200:
            data = resp.json()
            best = data.get("best_oa_location") or {}
            return best.get("url_for_pdf", "") or ""
    except Exception:
        pass
    return ""


needs_url = df_meta[(df_meta["open_access_url"].fillna("") == "") & (df_meta["doi"] != "")].copy()
print(f"OA URL 없는 논문: {len(needs_url)}편 → Unpaywall 조회")

upw_count = 0
for idx, row in tqdm(needs_url.iterrows(), total=len(needs_url), desc="Unpaywall"):
    url = get_unpaywall_pdf(row["doi"], UNPAYWALL_EMAIL)
    if url:
        df_meta.at[idx, "open_access_url"] = url
        upw_count += 1
    time.sleep(0.5)

print(f"Unpaywall 추가 확보: {upw_count}편")
has_url = (df_meta["open_access_url"].fillna("") != "").sum()
print(f"PDF URL 확보 합계: {has_url}편 / {len(df_meta)}편")
df_meta.to_excel(meta_path, index=False)


# %% [셀 9] Sci-Hub PDF 수집 (Unpaywall에서 못 얻은 논문)
"""
Sci-Hub는 유료 논문을 자동으로 제공하는 사이트입니다.
회사/기관 보안정책에 따라 접근이 차단될 수 있습니다.
SCIHUB_ENABLED = False로 설정하면 이 셀은 건너뜁니다.

주의: 논문 저작권에 관한 현지 법률과 기관 정책을 확인하세요.
"""
from bs4 import BeautifulSoup

HEADERS_BROWSER = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}


def _try_scihub_url(base_url: str, doi: str) -> str:
    """Sci-Hub 페이지에서 PDF 다운로드 URL을 파싱합니다."""
    target = f"{base_url.rstrip('/')}/{doi}"
    try:
        resp = requests.get(target, headers=HEADERS_BROWSER, timeout=20, allow_redirects=True)
        if resp.status_code != 200:
            return ""
        soup = BeautifulSoup(resp.text, "lxml")

        # 패턴 1: <iframe id="pdf" src="...">
        iframe = soup.find("iframe", id="pdf")
        if iframe and iframe.get("src"):
            src = iframe["src"]
            if src.startswith("//"):
                src = "https:" + src
            return src

        # 패턴 2: <embed src="..." type="application/pdf">
        embed = soup.find("embed", attrs={"type": "application/pdf"})
        if embed and embed.get("src"):
            src = embed["src"]
            if src.startswith("//"):
                src = "https:" + src
            return src

        # 패턴 3: 버튼/링크 텍스트에 save/download
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if href.lower().endswith(".pdf") or "pdf" in href.lower():
                if href.startswith("//"):
                    href = "https:" + href
                elif href.startswith("/"):
                    href = base_url.rstrip("/") + href
                return href

    except Exception:
        pass
    return ""


def fetch_pdf_url_scihub(doi: str) -> str:
    """여러 Sci-Hub 미러를 순서대로 시도해 PDF URL을 반환합니다."""
    for base_url in SCIHUB_URLS:
        url = _try_scihub_url(base_url, doi)
        if url:
            return url
        time.sleep(0.3)
    return ""


def download_pdf(url: str, save_path: str) -> bool:
    """URL에서 PDF를 다운로드합니다. 성공 여부를 반환합니다."""
    try:
        resp = requests.get(url, headers=HEADERS_BROWSER, timeout=30, stream=True)
        content_type = resp.headers.get("Content-Type", "")
        if resp.status_code == 200 and ("pdf" in content_type.lower() or url.lower().endswith(".pdf")):
            with open(save_path, "wb") as f:
                for chunk in resp.iter_content(chunk_size=8192):
                    f.write(chunk)
            return os.path.getsize(save_path) > 1000   # 1KB 이상이면 유효
        return False
    except Exception:
        return False


def safe_filename(doi_or_id: str, title: str, max_len: int = 80) -> str:
    name = re.sub(r'[\\/:*?"<>|]', "_", (doi_or_id or title or "unknown"))
    return name.strip()[:max_len] + ".pdf"


if not SCIHUB_ENABLED:
    print("Sci-Hub 비활성화 (SCIHUB_ENABLED=False)")
else:
    # OA URL이 없고 DOI가 있는 논문만 대상
    scihub_targets = df_meta[
        (df_meta["open_access_url"].fillna("") == "") &
        (df_meta["doi"].fillna("") != "")
    ].copy()
    print(f"Sci-Hub 대상: {len(scihub_targets)}편")

    scihub_results = []
    for _, row in tqdm(scihub_targets.iterrows(), total=len(scihub_targets), desc="Sci-Hub"):
        doi = row["doi"]
        fname = safe_filename(doi.replace("/", "_"), row.get("title", ""))
        fpath = os.path.join(PDF_DIR, fname)

        if os.path.exists(fpath) and os.path.getsize(fpath) > 1000:
            df_meta.loc[df_meta["doi"] == doi, "open_access_url"] = "file://" + fpath
            scihub_results.append({"doi": doi, "status": "기존파일"})
            continue

        pdf_url = fetch_pdf_url_scihub(doi)
        if pdf_url:
            ok = download_pdf(pdf_url, fpath)
            if ok:
                df_meta.loc[df_meta["doi"] == doi, "open_access_url"] = pdf_url
                scihub_results.append({"doi": doi, "status": "성공", "url": pdf_url})
            else:
                scihub_results.append({"doi": doi, "status": "다운로드실패", "url": pdf_url})
        else:
            scihub_results.append({"doi": doi, "status": "URL없음"})

        time.sleep(2.0)  # Sci-Hub 부하 방지 (필수)

    sh_df = pd.DataFrame(scihub_results)
    sh_ok = (sh_df["status"].isin(["성공", "기존파일"])).sum() if not sh_df.empty else 0
    print(f"Sci-Hub 성공: {sh_ok}편 / {len(scihub_targets)}편")
    df_meta.to_excel(meta_path, index=False)


# %% [셀 10] PDF 다운로드 (OA URL 기반)
"""
open_access_url이 있는 논문의 PDF를 pdf/ 폴더에 저장합니다.
"""

download_log = []
pdf_candidates = df_meta[df_meta["open_access_url"].fillna("") != ""].copy()
print(f"PDF 다운로드 대상: {len(pdf_candidates)}편")

for _, row in tqdm(pdf_candidates.iterrows(), total=len(pdf_candidates), desc="PDF 다운로드"):
    pid   = row["paper_id"]
    url   = row["open_access_url"]
    fname = safe_filename(str(pid).replace("/", "_"), row.get("title", ""))
    fpath = os.path.join(PDF_DIR, fname)

    if url.startswith("file://"):
        local = url[7:]
        if os.path.exists(local):
            download_log.append({"paper_id": pid, "pdf_path": local, "status": "로컬파일"})
            continue

    if os.path.exists(fpath) and os.path.getsize(fpath) > 1000:
        download_log.append({"paper_id": pid, "pdf_path": fpath, "status": "기존파일"})
        continue

    ok = download_pdf(url, fpath)
    if ok:
        download_log.append({"paper_id": pid, "pdf_path": fpath, "status": "성공"})
    else:
        download_log.append({"paper_id": pid, "pdf_path": "", "status": "실패"})

    time.sleep(0.3)

df_dl = pd.DataFrame(download_log) if download_log else pd.DataFrame(
    columns=["paper_id", "pdf_path", "status"]
)
ok_count = df_dl["status"].isin(["성공", "기존파일", "로컬파일"]).sum()
print(f"PDF 확보: {ok_count}편 / {len(pdf_candidates)}편")

df_meta = df_meta.merge(
    df_dl[["paper_id", "pdf_path"]].rename(columns={"pdf_path": "local_pdf_path"}),
    on="paper_id", how="left"
)
df_meta["local_pdf_path"] = df_meta["local_pdf_path"].fillna("")
df_meta.to_excel(meta_path, index=False)


# %% [셀 11] PDF 텍스트 추출
"""
pdfplumber → PyMuPDF 순으로 텍스트를 추출합니다.
추출 텍스트는 text/ 폴더에 저장됩니다.
"""
import pdfplumber

try:
    import fitz as pymupdf
    PYMUPDF_AVAILABLE = True
except ImportError:
    PYMUPDF_AVAILABLE = False
    print("PyMuPDF 없음 → pdfplumber만 사용")


def extract_text_from_pdf(pdf_path: str) -> str:
    # 1차: pdfplumber
    try:
        pages = []
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    pages.append(t)
        if pages:
            return "\n\n".join(pages)
    except Exception:
        pass

    # 2차: PyMuPDF
    if PYMUPDF_AVAILABLE:
        try:
            doc = pymupdf.open(pdf_path)
            pages = [page.get_text() for page in doc]
            doc.close()
            text = "\n\n".join(pages)
            if text.strip():
                return text
        except Exception:
            pass

    return ""


text_log = []
pdf_files = df_meta[df_meta["local_pdf_path"].fillna("") != ""].copy()
print(f"텍스트 추출 대상: {len(pdf_files)}편")

for _, row in tqdm(pdf_files.iterrows(), total=len(pdf_files), desc="텍스트 추출"):
    pid      = row["paper_id"]
    pdf_path = row["local_pdf_path"]
    if not os.path.exists(pdf_path):
        text_log.append({"paper_id": pid, "text_path": "", "text_len": 0, "status": "PDF없음"})
        continue

    txt_name = os.path.basename(pdf_path).replace(".pdf", ".txt")
    txt_path = os.path.join(TEXT_DIR, txt_name)

    if os.path.exists(txt_path):
        with open(txt_path, encoding="utf-8") as f:
            content = f.read()
        text_log.append({"paper_id": pid, "text_path": txt_path,
                          "text_len": len(content), "status": "기존파일"})
        continue

    text = extract_text_from_pdf(pdf_path)
    if text.strip():
        with open(txt_path, "w", encoding="utf-8") as f:
            f.write(text)
        text_log.append({"paper_id": pid, "text_path": txt_path,
                          "text_len": len(text), "status": "성공"})
    else:
        text_log.append({"paper_id": pid, "text_path": "", "text_len": 0, "status": "추출실패"})

df_txt = pd.DataFrame(text_log) if text_log else pd.DataFrame(
    columns=["paper_id", "text_path", "text_len", "status"]
)
ok_txt = df_txt["status"].isin(["성공", "기존파일"]).sum()
print(f"텍스트 추출 성공: {ok_txt}편")

df_meta = df_meta.merge(
    df_txt[["paper_id", "text_path", "text_len"]],
    on="paper_id", how="left"
)
df_meta["text_path"] = df_meta["text_path"].fillna("")
df_meta["text_len"]  = df_meta["text_len"].fillna(0).astype(int)
df_meta.to_excel(meta_path, index=False)


# %% [셀 12] 규칙 기반 합성 조건 추출
"""
src/ 모듈의 패턴 매칭으로 합성 조건을 추출합니다.
초록(abstract)과 본문(text_path)에 모두 적용합니다.
"""
import sys as _sys
if BASE_DIR not in _sys.path:
    _sys.path.insert(0, BASE_DIR)

RULE_MODULES_OK = all(
    os.path.exists(os.path.join(BASE_DIR, "src", f))
    for f in ["experiment_parser.py", "quantity_extractor.py",
              "extract_ceria_rules.py", "ceria_dictionary.py", "dopant_dictionary.py"]
)
print(f"규칙 기반 모듈 준비: {RULE_MODULES_OK}")

rule_records = []

if RULE_MODULES_OK:
    from src.experiment_parser import parse_experiments_from_text
    from src.extract_ceria_rules import extract_ceria_fields_from_text

    for _, row in tqdm(df_meta.iterrows(), total=len(df_meta), desc="규칙 기반 추출"):
        pid = row["paper_id"]

        # 텍스트 소스 결정: 본문 > 초록
        text = ""
        src_label = "abstract"
        if row.get("text_path") and os.path.exists(str(row.get("text_path", ""))):
            with open(row["text_path"], encoding="utf-8") as f:
                text = f.read()
            src_label = "fulltext"
        elif row.get("abstract"):
            text = str(row["abstract"])

        if not text.strip():
            continue

        # 실험 블록 파싱
        try:
            blocks = parse_experiments_from_text(text, paper_id=pid)
        except Exception as e:
            print(f"  {str(pid)[:20]}: parser 오류 - {e}")
            blocks = []

        if blocks:
            for rec in blocks:
                rec["paper_id"]    = pid
                rec["text_source"] = src_label
                rule_records.append(rec)
        else:
            # 블록 미검출 → 전체 텍스트에서 직접 ceria fields 추출
            try:
                fields = extract_ceria_fields_from_text(text)
                fields["paper_id"]    = pid
                fields["text_source"] = src_label
                fields["experiment_id"] = f"{pid}_EXP001"
                rule_records.append(fields)
            except Exception:
                pass

    print(f"규칙 기반 추출: {len(rule_records)}개 실험 블록")
else:
    print("src/ 모듈 준비 안 됨 → 셀 13 Claude 추출로 진행")


# %% [셀 13] LLM API 기반 추출 (OpenAI 또는 Anthropic)
"""
OPENAI_API_KEY 또는 ANTHROPIC_API_KEY가 있을 때 규칙 기반에서 못 얻은 정보를 보완합니다.
OpenAI: gpt-4o-mini (빠르고 저렴, 약 $0.0003/편)
Anthropic: claude-haiku-4-5 (약 $0.0015/편)
"""

claude_records = []  # OpenAI/Claude 추출 결과 모두 이 변수에 저장
_LLM_CACHE_PATH = os.path.join(BASE_DIR, "output", "llm_cache.json")

EXTRACTION_PROMPT = """You are an expert materials scientist specializing in CeO2 (ceria) nanoparticle synthesis. Extract synthesis conditions from the paper text provided.

Return ONLY a valid JSON object with these exact fields. Use null for missing or unclear information. Numeric fields must be plain numbers only (no units, no text).

=== SYNTHESIS METHOD INFERENCE RULES ===
Even if not explicitly named, infer from context:
- "autoclave", "Teflon-lined vessel" → hydrothermal
- "citric acid"+"gel", "Pechini", "alkoxide" → sol-gel
- "co-precipitation", "NaOH/NH3 added dropwise"+"pH" → co-precipitation
- "NaOH/NH3"+(no autoclave) → precipitation
- "glycine fuel", "urea fuel", "self-ignition" → combustion
- "microwave irradiation" → microwave
- "ball mill", "mechanical grinding" → mechanochemical
- "sonication" during synthesis → sonochemical
If uncertain, use "wet chemical" rather than null.

Fields:
- synthesis_method: hydrothermal/solvothermal/sol-gel/precipitation/co-precipitation/combustion/spray_pyrolysis/microwave/template/thermal_decomposition/mechanochemical/sonochemical/wet chemical/other
- ce_precursor: chemical formula preferred (Ce(NO3)3·6H2O, CeCl3·7H2O, Ce(CH3COO)3, (NH4)2Ce(NO3)6, Ce(acac)3)
- ce_precursor_amount: e.g. "2 mmol", "0.868 g"
- precursor_concentration: molar concentration if stated, e.g. "0.1 M", "0.5 mol/L" (text)
- solvent: e.g. "distilled water", "ethanol", "water/ethanol 1:1"
- solvent_amount: e.g. "40 mL"
- additive: pH agent, surfactant, or complexing agent
- additive_amount: e.g. "4 mmol"
- synthesis_temperature_c: NUMBER °C (reaction/hydrothermal temp, NOT calcination)
- synthesis_time_h: NUMBER hours (overnight→12, 1 day→24, 30 min→0.5)
- ph_synthesis: NUMBER — pH value during synthesis/precipitation (e.g. 10, 9.5); null if not stated
- atmosphere: air/N2/Ar/O2/vacuum/autoclave/other
- calcination_temperature_c: NUMBER °C
- calcination_time_h: NUMBER hours
- drying_temperature_c: NUMBER °C
- particle_size_tem_nm: NUMBER from TEM only
- particle_size_sem_nm: NUMBER from SEM only
- crystallite_size_xrd_nm: NUMBER from XRD Scherrer only
- morphology: sphere/cube/rod/wire/flower/octahedron/plate/porous/hollow/other
- crystal_phase: e.g. "fluorite cubic", "Ce2O3", "amorphous", "mixed"
- dopant: element symbol only (e.g. "Sm", "Gd"), null if pure CeO2
- dopant_concentration: dopant amount as text (e.g. "5 mol%", "10 at%", "x=0.1"); null if undoped
- dopant_formula: full Ce-dopant formula if stated (e.g. "Ce0.9Sm0.1O2-δ", "Ce0.8Zr0.2O2"); null if not stated
- bet_surface_area: NUMBER m²/g
- notes: one sentence on key synthesis details

--- EXAMPLES ---

Example 1 (Hydrothermal, undoped):
Text: "Ce(NO3)3·6H2O (0.868 g, 2 mmol) was dissolved in 40 mL distilled water. NaOH solution (4 mmol) was added dropwise with stirring. The suspension was transferred to a Teflon-lined autoclave and maintained at 180°C for 24 h. After cooling, the product was washed with water and ethanol, dried at 80°C overnight, and calcined at 500°C for 2 h in air. XRD confirmed fluorite cubic structure with crystallite size of 9.3 nm."
Output: {"synthesis_method": "hydrothermal", "ce_precursor": "Ce(NO3)3·6H2O", "ce_precursor_amount": "2 mmol", "precursor_concentration": null, "solvent": "distilled water", "solvent_amount": "40 mL", "additive": "NaOH", "additive_amount": "4 mmol", "synthesis_temperature_c": 180, "synthesis_time_h": 24, "ph_synthesis": null, "atmosphere": "autoclave", "calcination_temperature_c": 500, "calcination_time_h": 2, "drying_temperature_c": 80, "particle_size_tem_nm": null, "particle_size_sem_nm": null, "crystallite_size_xrd_nm": 9.3, "morphology": null, "crystal_phase": "fluorite cubic", "dopant": null, "dopant_concentration": null, "dopant_formula": null, "bet_surface_area": null, "notes": "NaOH added dropwise before hydrothermal treatment"}

Example 2 (Sol-gel, undoped):
Text: "Cerium(III) nitrate hexahydrate (4.34 g) and citric acid (6.30 g, Ce:citric = 1:1.5 mol) were dissolved in 100 mL deionized water. The solution was heated at 80°C with stirring until a transparent gel formed (~3 h). The gel was dried at 120°C overnight, then calcined at 600°C for 4 h in air. TEM showed spherical particles with mean diameter 8 nm. BET surface area was 62 m²/g."
Output: {"synthesis_method": "sol-gel", "ce_precursor": "Ce(NO3)3·6H2O", "ce_precursor_amount": "4.34 g", "precursor_concentration": null, "solvent": "deionized water", "solvent_amount": "100 mL", "additive": "citric acid", "additive_amount": "6.30 g", "synthesis_temperature_c": 80, "synthesis_time_h": 3, "ph_synthesis": null, "atmosphere": "air", "calcination_temperature_c": 600, "calcination_time_h": 4, "drying_temperature_c": 120, "particle_size_tem_nm": 8, "particle_size_sem_nm": null, "crystallite_size_xrd_nm": null, "morphology": "sphere", "crystal_phase": null, "dopant": null, "dopant_concentration": null, "dopant_formula": null, "bet_surface_area": 62, "notes": "Ce:citric acid molar ratio 1:1.5; gelation at 80°C"}

Example 3 (Co-precipitation, doped):
Text: "CeCl3·7H2O (1.35 mmol) and SmCl3·6H2O (0.15 mmol) were dissolved in 50 mL ethanol. 25% ammonia solution was added dropwise until pH reached 10. The precipitate was aged at room temperature for 12 h, filtered, and calcined at 400°C for 3 h. Scherrer analysis yielded crystallite size 5.2 nm. The composition was Ce0.9Sm0.1O2-δ. SEM revealed rod-like morphology."
Output: {"synthesis_method": "co-precipitation", "ce_precursor": "CeCl3·7H2O", "ce_precursor_amount": "1.35 mmol", "precursor_concentration": null, "solvent": "ethanol", "solvent_amount": "50 mL", "additive": "ammonia solution (25%)", "additive_amount": null, "synthesis_temperature_c": null, "synthesis_time_h": 12, "ph_synthesis": 10, "atmosphere": "air", "calcination_temperature_c": 400, "calcination_time_h": 3, "drying_temperature_c": null, "particle_size_tem_nm": null, "particle_size_sem_nm": null, "crystallite_size_xrd_nm": 5.2, "morphology": "rod", "crystal_phase": null, "dopant": "Sm", "dopant_concentration": "10 mol%", "dopant_formula": "Ce0.9Sm0.1O2-δ", "bet_surface_area": null, "notes": "pH adjusted to 10; aged 12 h before filtration"}

Example 4 (Hydrothermal, Zr-doped):
Text: "Ce(NO3)3·6H2O (0.1 M) and ZrO(NO3)2·xH2O were mixed in water with Ce:Zr = 8:2 molar ratio to prepare Ce0.8Zr0.2O2. The solution pH was adjusted to 9 with NaOH. After hydrothermal treatment at 160°C for 12 h, the product was calcined at 550°C for 3 h. BET surface area: 89 m²/g. TEM showed particles of ~7 nm."
Output: {"synthesis_method": "hydrothermal", "ce_precursor": "Ce(NO3)3·6H2O", "ce_precursor_amount": null, "precursor_concentration": "0.1 M", "solvent": "water", "solvent_amount": null, "additive": "NaOH", "additive_amount": null, "synthesis_temperature_c": 160, "synthesis_time_h": 12, "ph_synthesis": 9, "atmosphere": "autoclave", "calcination_temperature_c": 550, "calcination_time_h": 3, "drying_temperature_c": null, "particle_size_tem_nm": 7, "particle_size_sem_nm": null, "crystallite_size_xrd_nm": null, "morphology": null, "crystal_phase": null, "dopant": "Zr", "dopant_concentration": "20 mol%", "dopant_formula": "Ce0.8Zr0.2O2", "bet_surface_area": 89, "notes": "Ce:Zr = 8:2 molar ratio; NaOH pH adjustment to 9"}

--- NOW EXTRACT ---

Text:
"""

def _extract_experimental_section(text: str, max_chars: int = 6000) -> str:
    """Experimental/Synthesis 섹션을 우선 추출. 없으면 text[:max_chars] 반환."""
    section_re = re.compile(
        r'\n[ \t]*(?:\d+[\.\d]*\s*)?'
        r'(?:experimental(?:\s+(?:section|details?|procedure|part|methods?))?'
        r'|materials?\s+and\s+methods?'
        r'|synthesis(?:\s+of\s+[\w\s]{0,30})?'
        r'|preparation(?:\s+of\s+[\w\s]{0,30})?'
        r'|sample\s+preparation'
        r'|nanoparticle\s+synthesis)'
        r'\s*[\n:]',
        re.IGNORECASE
    )
    end_re = re.compile(
        r'\n[ \t]*(?:\d+[\.\d]*\s*)?'
        r'(?:results?(?:\s+and\s+discussion)?|discussion|characterization|'
        r'conclusions?|references?|acknowledgements?|supporting\s+information)'
        r'\s*\n',
        re.IGNORECASE
    )
    m = section_re.search(text)
    if m:
        start = m.end()
        end_m = end_re.search(text, start)
        end = end_m.start() if end_m else start + max_chars
        snippet = text[start:end].strip()
        if len(snippet) > 300:
            return snippet[:max_chars]
    return text[:max_chars]

def _build_llm_targets(df_meta, rule_records):
    # Target ALL papers with text — LLM adds ce_precursor/solvent/additive
    # that rule-based regex cannot extract, regardless of prior rule coverage
    targets = []
    for _, row in df_meta.iterrows():
        pid = row["paper_id"]
        text = ""
        if row.get("text_path") and os.path.exists(str(row.get("text_path", ""))):
            with open(row["text_path"], encoding="utf-8") as f:
                text = f.read()
        elif row.get("abstract"):
            text = str(row["abstract"])
        if text.strip():
            targets.append((pid, row.get("title", ""), text))
    return targets

def _parse_llm_response(content: str, paper_id: str, title: str):
    m = re.search(r"\{.*\}", content, re.DOTALL)
    if m:
        try:
            extracted = json.loads(m.group())
            extracted["paper_id"] = paper_id
            extracted["title"]    = title
            return extracted
        except json.JSONDecodeError:
            pass
    return None


# ── OpenAI import ────────────────────────────────────────────────────────────
if OPENAI_API_KEY:
    try:
        from openai import OpenAI as _OpenAI
    except ImportError:
        print("openai 패키지 없음 → Anaconda Prompt에서: pip install openai")
        OPENAI_API_KEY = ""

# ── 재실행 방지 / 캐시 로드 ────────────────────────────────────────────────────
if globals().get('_llm_extraction_done', False):
    print(f"✓ 이미 추출 완료 ({len(claude_records)}편) → 재실행 건너뜀. 셀 14를 실행하세요.")

elif os.path.exists(_LLM_CACHE_PATH):
    with open(_LLM_CACHE_PATH, encoding="utf-8") as _f:
        claude_records = json.load(_f)
    globals()['_llm_extraction_done'] = True
    print(f"✓ 캐시 로드 완료: {len(claude_records)}편 (output/llm_cache.json) → 셀 14를 실행하세요.")

# ── OpenAI 추출 ───────────────────────────────────────────────────────────────
elif OPENAI_API_KEY:
    _oa_client = _OpenAI(api_key=OPENAI_API_KEY)

    def _extract_with_openai(text: str, paper_id: str, title: str):
        snippet = _extract_experimental_section(text, max_chars=6000)
        try:
            resp = _oa_client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[{"role": "user", "content": EXTRACTION_PROMPT + snippet}],
                max_tokens=1200,
                temperature=0,
                response_format={"type": "json_object"},
            )
            content = resp.choices[0].message.content
            return _parse_llm_response(content, paper_id, title)
        except Exception as e:
            print(f"  OpenAI 오류 ({str(paper_id)[:20]}): {e}")
        return None

    llm_targets = _build_llm_targets(df_meta, rule_records)
    print(f"OpenAI 추출 대상: {len(llm_targets)}편")
    approx_cost = len(llm_targets) * 0.0003
    print(f"예상 비용: 약 ${approx_cost:.2f} (gpt-4o-mini 기준)")

    for pid, title, text in tqdm(llm_targets, desc="OpenAI 추출"):
        result = _extract_with_openai(text, pid, title)
        if result:
            claude_records.append(result)
        time.sleep(0.05)

    print(f"OpenAI 추출 성공: {len(claude_records)}편")
    with open(_LLM_CACHE_PATH, "w", encoding="utf-8") as _f:
        json.dump(claude_records, _f, ensure_ascii=False)
    print(f"✓ 캐시 저장: {_LLM_CACHE_PATH}")
    globals()['_llm_extraction_done'] = True

# ── Anthropic 사용 (OpenAI 키 없을 때) ───────────────────────────────────────
elif ANTHROPIC_API_KEY:
    import anthropic as _anthropic
    _ant_client = _anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

    def _extract_with_claude(text: str, paper_id: str, title: str):
        snippet = _extract_experimental_section(text, max_chars=6000)
        try:
            msg = _ant_client.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=1200,
                messages=[{"role": "user", "content": EXTRACTION_PROMPT + snippet}],
            )
            content = msg.content[0].text
            return _parse_llm_response(content, paper_id, title)
        except Exception as e:
            print(f"  Claude 오류 ({str(paper_id)[:20]}): {e}")
        return None

    llm_targets = _build_llm_targets(df_meta, rule_records)
    print(f"Claude 추출 대상: {len(llm_targets)}편")
    for pid, title, text in tqdm(llm_targets, desc="Claude 추출"):
        result = _extract_with_claude(text, pid, title)
        if result:
            claude_records.append(result)
        time.sleep(0.2)
    print(f"Claude 추출 성공: {len(claude_records)}편")
    with open(_LLM_CACHE_PATH, "w", encoding="utf-8") as _f:
        json.dump(claude_records, _f, ensure_ascii=False)
    print(f"✓ 캐시 저장: {_LLM_CACHE_PATH}")
    globals()['_llm_extraction_done'] = True

else:
    print("API 키 없음 → LLM 추출 건너뜁니다. (.env 파일에 OPENAI_API_KEY 입력 후 재실행)")


# %% [셀 14] 결과 통합 및 Excel 출력
"""
규칙 기반 + Claude 결과를 병합하여 정돈된 Excel로 저장합니다.
커널 재시작 후에도 저장된 파일(papers_metadata.xlsx, llm_cache.json)에서
자동으로 로드하여 단독 실행 가능합니다.
"""
import os, json
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 커널 재시작 시 저장된 파일에서 자동 로드 ──────────────────────────────────
try:
    BASE_DIR
except NameError:
    BASE_DIR = r"d:\머신러닝 교육\ceria_pipeline_data"

OUTPUT_DIR = os.path.join(BASE_DIR, "output")

# df_meta: 메타데이터 Excel에서 로드
if 'df_meta' not in dir() or df_meta is None:
    _meta_path = os.path.join(OUTPUT_DIR, "papers_metadata.xlsx")
    df_meta = pd.read_excel(_meta_path)
    print(f"[자동로드] papers_metadata.xlsx ({len(df_meta)}편)")

# claude_records: LLM 캐시에서 로드
if 'claude_records' not in dir() or not claude_records:
    _cache_path = os.path.join(OUTPUT_DIR, "llm_cache.json")
    if os.path.exists(_cache_path):
        with open(_cache_path, encoding="utf-8") as _f:
            claude_records = json.load(_f)
        print(f"[자동로드] llm_cache.json ({len(claude_records)}편)")
    else:
        claude_records = []

# rule_records: 규칙 기반 결과 없으면 빈 리스트
if 'rule_records' not in dir():
    rule_records = []

SYNTH_COLS = [
    "synthesis_method", "ce_precursor", "ce_precursor_amount",
    "precursor_concentration",                           # 신규: 전구체 몰농도
    "solvent", "solvent_amount", "additive", "additive_amount",
    "synthesis_temperature_c", "synthesis_time_h",
    "ph_synthesis",                                      # 신규: 합성 pH
    "atmosphere",
    "calcination_temperature_c", "calcination_time_h", "drying_temperature_c",
    "particle_size_tem_nm", "particle_size_sem_nm", "crystallite_size_xrd_nm",
    "morphology", "crystal_phase",
    "dopant", "dopant_concentration", "dopant_formula",  # 신규: 도핑 농도/화학식
    "bet_surface_area",
    "notes",
]

META_COLS = [
    "paper_id", "title", "authors", "year", "pub_date",
    "journal", "doi", "citation_count", "source_api",
    "abstract", "open_access_url", "local_pdf_path", "text_source",
]

# ── 규칙 기반 결과: paper_id당 첫 번째 블록만 사용 ──────────────────────────
df_rule_raw = pd.DataFrame(rule_records) if rule_records else pd.DataFrame()
df_rule_norm = pd.DataFrame()
if not df_rule_raw.empty:
    available = [c for c in SYNTH_COLS if c in df_rule_raw.columns]
    df_rule_norm = (
        df_rule_raw[["paper_id"] + available]
        .copy()
        .drop_duplicates(subset=["paper_id"], keep="first")
    )
    df_rule_norm["extraction_source"] = "rule_based"

# ── Claude 결과 ───────────────────────────────────────────────────────────────
df_claude_norm = pd.DataFrame()
if claude_records:
    df_cl = pd.DataFrame(claude_records)
    available = [c for c in SYNTH_COLS if c in df_cl.columns]
    df_claude_norm = df_cl[["paper_id"] + available].copy()
    df_claude_norm["extraction_source"] = "claude_api"

# ── 통합: 규칙 기반 유지 + LLM 전용 컬럼(ce_precursor 등) 보완 ─────────────
# LLM만 추출 가능한 컬럼 (규칙 기반에는 없음)
_LLM_ONLY = ["ce_precursor", "ce_precursor_amount",
             "solvent", "solvent_amount",
             "additive", "additive_amount",
             "dopant", "notes"]

if not df_rule_norm.empty and not df_claude_norm.empty:
    # 규칙 기반 레코드에 LLM 전용 컬럼만 merge하여 보완
    _fill_cols = ["paper_id"] + [c for c in _LLM_ONLY if c in df_claude_norm.columns]
    _claude_fill = df_claude_norm[_fill_cols]
    df_rule_aug = df_rule_norm.merge(_claude_fill, on="paper_id", how="left")
    # LLM만 처리된 논문(규칙 기반 없음) 추가
    _rule_pids = set(df_rule_norm["paper_id"])
    _claude_only = df_claude_norm[~df_claude_norm["paper_id"].isin(_rule_pids)]
    df_synth = pd.concat([df_rule_aug, _claude_only], ignore_index=True)
elif not df_rule_norm.empty:
    df_synth = df_rule_norm
elif not df_claude_norm.empty:
    df_synth = df_claude_norm
else:
    df_synth = pd.DataFrame({"paper_id": df_meta["paper_id"]})

# ── 메타데이터와 병합 ─────────────────────────────────────────────────────────
meta_for_merge = df_meta[[c for c in META_COLS if c in df_meta.columns]].copy()
df_final = meta_for_merge.merge(df_synth, on="paper_id", how="left")

# 컬럼 순서 정렬
ordered_cols = [c for c in META_COLS if c in df_final.columns]
ordered_cols += [c for c in SYNTH_COLS if c in df_final.columns and c not in ordered_cols]
ordered_cols += ["extraction_source"]
df_final = df_final[[c for c in ordered_cols if c in df_final.columns]]

# ── Excel 작성 ─────────────────────────────────────────────────────────────────
output_path = os.path.join(OUTPUT_DIR, "ceria_synthesis_database.xlsx")

HEADER_FILL   = PatternFill("solid", fgColor="2F5496")  # 진파랑
META_FILL     = PatternFill("solid", fgColor="D9E2F3")  # 연파랑
SYNTH_FILL    = PatternFill("solid", fgColor="E2EFDA")  # 연초록
HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
BODY_FONT     = Font(name="Calibri", size=9)
CENTER_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=False)
WRAP_ALIGN    = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
THIN_BORDER   = Border(
    left=Side(style="thin", color="BFBFBF"), right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),  bottom=Side(style="thin", color="BFBFBF"),
)

META_HEADER_FILL  = PatternFill("solid", fgColor="4472C4")
SYNTH_HEADER_FILL = PatternFill("solid", fgColor="375623")

SYNTH_COL_SET = set(SYNTH_COLS)

def _style_sheet(ws, df, freeze_col: int = 1):
    """공통 헤더 스타일 + 열 너비 적용."""
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        if col_name in SYNTH_COL_SET:
            cell.fill = SYNTH_HEADER_FILL
        else:
            cell.fill = META_HEADER_FILL

    # 열 너비 자동
    col_widths = {
        "title": 50, "abstract": 60, "authors": 35, "notes": 40,
        "ce_precursor": 30, "solvent": 25, "additive": 30,
        "chemical_quantity_pairs": 50, "raw_block_text": 60,
    }
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        width = col_widths.get(col_name, 15)
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = ws.cell(row=2, column=freeze_col + 1)
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 20


def _clean_for_excel(val):
    """Excel에 쓸 수 없는 제어문자 및 특수문자 제거, 32000자 초과 시 자름."""
    if not isinstance(val, str):
        return val
    # openpyxl 불허 문자 제거 (U+0000~U+001F 중 탭·개행 제외, U+FFFE, U+FFFF)
    val = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f￾￿]', ' ', val)
    # 서로게이트 문자 제거
    val = re.sub(r'[\ud800-\udfff]', '', val)
    # Excel 수식 오인 방지: 셀 값이 =로 시작하면 앞에 공백 추가
    if val.startswith('=') or val.startswith('+') or val.startswith('-') or val.startswith('@'):
        val = ' ' + val
    return val[:32000] if len(val) > 32000 else val

def _clean_df(df):
    return df.apply(lambda col: col.map(lambda x: _clean_for_excel(x) if isinstance(x, str) else x))

with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    _clean_df(df_final).to_excel(writer, sheet_name="합성조건", index=False)
    _clean_df(df_meta).to_excel(writer, sheet_name="논문목록_전체", index=False)
    if not df_rule_raw.empty:
        _clean_df(df_rule_raw).to_excel(writer, sheet_name="규칙기반_상세", index=False)
    if claude_records:
        _clean_df(pd.DataFrame(claude_records)).to_excel(writer, sheet_name="Claude_추출", index=False)

    wb = writer.book
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        src_df = {
            "합성조건": df_final,
            "논문목록_전체": df_meta,
        }.get(sheet_name)
        if src_df is not None:
            _style_sheet(ws, src_df)

print(f"\n최종 Excel 저장: {output_path}")
print(f"총 {len(df_final)}편 | 합성조건 추출: {df_final['synthesis_method'].notna().sum()}편")


# %% [셀 15] 요약 통계
import pandas as pd, os
_output_path = os.path.join(os.path.dirname(os.path.abspath(__file__))
                            if '__file__' in globals() else os.getcwd(),
                            "output", "ceria_synthesis_database.xlsx")
if 'df_final' not in dir() or df_final is None:
    df_final = pd.read_excel(_output_path, sheet_name="합성조건")
    print(f"[Excel에서 로드] {_output_path}")

print("=" * 65)
print("수집 완료 요약")
print("=" * 65)
print(f"총 논문수:        {len(df_final):,}편")
print(f"연도 범위:        {df_final['year'].min()} ~ {df_final['year'].max()}")
print(f"  DOI 확보:       {(df_final['doi'].fillna('') != '').sum():,}편")
print(f"  PDF 확보:       {(df_final['local_pdf_path'].fillna('') != '').sum():,}편")
print(f"  합성조건 추출:  {df_final['synthesis_method'].notna().sum():,}편")
print(f"  추출소스 - 규칙기반: {(df_final['extraction_source'] == 'rule_based').sum():,}편")
print(f"  추출소스 - Claude:   {(df_final['extraction_source'] == 'claude_api').sum():,}편")

if "synthesis_method" in df_final.columns:
    sm = df_final["synthesis_method"].dropna()
    if len(sm) > 0:
        # 여러 방법이 ; 로 구분될 수 있으므로 분리
        all_methods = []
        for val in sm:
            all_methods.extend([v.strip() for v in str(val).split(";")])
        method_counts = pd.Series(all_methods).value_counts()
        print("\n합성 방법 분포 (상위 10):")
        for method, cnt in method_counts.head(10).items():
            print(f"  {method:<30}: {cnt:,}편")

if "particle_size_tem_nm" in df_final.columns:
    sizes = pd.to_numeric(df_final["particle_size_tem_nm"], errors="coerce").dropna()
    if len(sizes) > 0:
        print(f"\nTEM 입자크기 (nm): 평균 {sizes.mean():.1f}, 중앙값 {sizes.median():.1f}, "
              f"범위 {sizes.min():.1f}~{sizes.max():.1f} (n={len(sizes)})")

print(f"\n출력 파일: {_output_path}")


# %% [셀 16] 컬럼별 채움률 분석
import pandas as pd, os
_output_path = os.path.join(os.path.dirname(os.path.abspath(__file__))
                            if '__file__' in globals() else os.getcwd(),
                            "output", "ceria_synthesis_database.xlsx")
if 'df_final' not in dir() or df_final is None:
    df_final = pd.read_excel(_output_path, sheet_name="합성조건")

SYNTH_COLS = [
    "synthesis_method", "ce_precursor", "ce_precursor_amount",
    "solvent", "solvent_amount", "additive", "additive_amount",
    "synthesis_temperature_c", "synthesis_time_h", "atmosphere",
    "calcination_temperature_c", "calcination_time_h", "drying_temperature_c",
    "particle_size_tem_nm", "particle_size_sem_nm", "crystallite_size_xrd_nm",
    "morphology", "crystal_phase", "dopant", "bet_surface_area", "notes",
]
total = len(df_final)
print(f"전체 논문: {total:,}편\n")
print(f"{'컬럼':<32} {'채움':>6} {'비율':>6}")
print("-" * 48)
for col in SYNTH_COLS:
    if col in df_final.columns:
        filled = df_final[col].replace("", pd.NA).notna().sum()
        print(f"  {col:<30} {filled:>5,}  {filled/total*100:>5.1f}%")

# 초록 있는 논문 중 합성방법 미추출 비율
has_abstract = (df_final["abstract"].fillna("") != "").sum() if "abstract" in df_final.columns else 0
no_method = df_final["synthesis_method"].isna().sum()
print(f"\n초록 보유: {has_abstract:,}편 | 합성방법 미추출: {no_method:,}편")
has_pdf = (df_final["local_pdf_path"].fillna("") != "").sum() if "local_pdf_path" in df_final.columns else 0
print(f"PDF 보유:  {has_pdf:,}편")


# %% [셀 17] 데이터 보완 - 키워드/정규식으로 빈 컬럼 채우기
import pandas as pd, os, re

_output_path = os.path.join(os.path.dirname(os.path.abspath(__file__))
                            if '__file__' in globals() else os.getcwd(),
                            "output", "ceria_synthesis_database.xlsx")
df = pd.read_excel(_output_path, sheet_name="합성조건")
total = len(df)
print(f"보완 시작: {total:,}편\n")

# ── 1. 합성방법 키워드 매칭 ──────────────────────────────────────────────────
# 순서 중요: 구체적인 방법을 먼저 검사
METHOD_KEYWORDS = [
    ("co-precipitation",      ["co-precipitation", "coprecipitation", "co precipitation"]),
    ("solvothermal",          ["solvothermal"]),
    ("hydrothermal",          ["hydrothermal"]),
    ("sol-gel",               ["sol-gel", "sol gel", "pechini", "citrate gel"]),
    ("combustion",            ["combustion", "auto-combustion", "autocombustion", "solution combustion"]),
    ("spray_pyrolysis",       ["spray pyrolysis", "spray-pyrolysis"]),
    ("microwave",             ["microwave"]),
    ("thermal_decomposition", ["thermal decomposition", "thermolysis", "thermally decomposed"]),
    ("mechanochemical",       ["mechanochemical", "ball mill", "ball-mill", "milling"]),
    ("sonochemical",          ["sonochemical", "sonication-assisted synthesis", "ultrasound synthesis"]),
    ("template",              ["hard template", "soft template", "template-assisted"]),
    ("precipitation",         ["precipitation"]),
]

def detect_method(text):
    if not text:
        return None
    t = text.lower()
    for method, keywords in METHOD_KEYWORDS:
        for kw in keywords:
            if kw in t:
                return method
    return None

filled_method = 0
for idx, row in df.iterrows():
    if pd.notna(row.get("synthesis_method")) and str(row.get("synthesis_method", "")).strip():
        continue
    combined = " ".join([str(row.get("title", "") or ""), str(row.get("abstract", "") or "")])
    method = detect_method(combined)
    if method:
        df.at[idx, "synthesis_method"] = method
        filled_method += 1
print(f"  합성방법 추가: {filled_method:,}편")

# ── 2. 형태(morphology) 키워드 매칭 ─────────────────────────────────────────
MORPHOLOGY_KEYWORDS = [
    ("rod",        ["nanorod", "nano-rod", "rod-shaped", "rod-like", "nanorods"]),
    ("wire",       ["nanowire", "nano-wire", "wire-like", "nanowires"]),
    ("cube",       ["nanocube", "nanocubes", "cube-shaped", "cubic nanoparticle"]),
    ("flower",     ["nanoflower", "flower-like", "flowerlike", "nanoflowers"]),
    ("plate",      ["nanoplate", "nanoplates", "plate-like", "platelet"]),
    ("octahedron", ["octahedral", "octahedron"]),
    ("hollow",     ["hollow sphere", "hollow nanoparticle", "core-shell"]),
    ("porous",     ["mesoporous", "nanoporous", "porous ceo2", "porous ceria"]),
    ("sphere",     ["nanosphere", "spherical nanoparticle", "quasi-spherical"]),
]

filled_morph = 0
for idx, row in df.iterrows():
    if pd.notna(row.get("morphology")) and str(row.get("morphology", "")).strip():
        continue
    combined = " ".join([str(row.get("title", "") or ""), str(row.get("abstract", "") or "")]).lower()
    for morph, keywords in MORPHOLOGY_KEYWORDS:
        if any(kw in combined for kw in keywords):
            df.at[idx, "morphology"] = morph
            filled_morph += 1
            break
print(f"  형태    추가: {filled_morph:,}편")

# ── 3. XRD 결정립 크기 정규식 추출 ──────────────────────────────────────────
XRD_PATTERNS = [
    r'(?:crystallite|crystalline)\s*size[^.]{0,80}?(\d+\.?\d*)\s*nm',
    r'scherrer[^.]{0,80}?(\d+\.?\d*)\s*nm',
    r'(?:xrd|x-ray)[^.]{0,80}?(\d+\.?\d*)\s*nm',
    r'(?:mean|average)\s+crystallite[^.]{0,60}?(\d+\.?\d*)\s*nm',
]
TEM_PATTERNS = [
    r'tem[^.]{0,80}?(\d+\.?\d*)\s*nm',
    r'transmission\s+electron\s+microscop[^.]{0,80}?(\d+\.?\d*)\s*nm',
    r'(?:mean|average)\s+(?:particle|grain)\s+size[^.]{0,60}?(\d+\.?\d*)\s*nm',
]
BET_PATTERNS = [
    r'bet[^.]{0,80}?(\d+\.?\d*)\s*m[²2][/\s]?g',
    r'surface\s+area[^.]{0,80}?(\d+\.?\d*)\s*m[²2][/\s]?g',
    r'(\d+\.?\d*)\s*m[²2]\s*/\s*g',
]

def _find_num(text, patterns, lo, hi):
    if not text:
        return None
    for pat in patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            try:
                val = float(m.group(1))
                if lo <= val <= hi:
                    return val
            except Exception:
                pass
    return None

filled_xrd = filled_tem = filled_bet = 0
for idx, row in df.iterrows():
    abstract = str(row.get("abstract", "") or "")
    if pd.isna(row.get("crystallite_size_xrd_nm")) or str(row.get("crystallite_size_xrd_nm", "")).strip() in ("", "nan"):
        v = _find_num(abstract, XRD_PATTERNS, 0.5, 200)
        if v:
            df.at[idx, "crystallite_size_xrd_nm"] = v
            filled_xrd += 1
    if pd.isna(row.get("particle_size_tem_nm")) or str(row.get("particle_size_tem_nm", "")).strip() in ("", "nan"):
        v = _find_num(abstract, TEM_PATTERNS, 0.5, 500)
        if v:
            df.at[idx, "particle_size_tem_nm"] = v
            filled_tem += 1
    if pd.isna(row.get("bet_surface_area")) or str(row.get("bet_surface_area", "")).strip() in ("", "nan"):
        v = _find_num(abstract, BET_PATTERNS, 1, 1000)
        if v:
            df.at[idx, "bet_surface_area"] = v
            filled_bet += 1

print(f"  XRD 결정립크기 추가: {filled_xrd:,}편")
print(f"  TEM 입자크기 추가:   {filled_tem:,}편")
print(f"  BET 표면적 추가:     {filled_bet:,}편")

# ── 4. 도핑 원소 키워드 매칭 ────────────────────────────────────────────────
DOPANTS = ["Sm", "Gd", "La", "Zr", "Y", "Pr", "Nd", "Eu", "Tb", "Dy",
           "Ho", "Er", "Yb", "Lu", "Ca", "Sr", "Ba", "Cu", "Fe", "Co",
           "Ni", "Mn", "Ti", "Al", "Si", "Mg"]
DOPANT_PATTERNS = [re.compile(
    rf'\b{d}(?:-doped|doped|/CeO2|[- ]doped\s+ceri|[- ]doped\s+ceo)', re.IGNORECASE
) for d in DOPANTS]

filled_dopant = 0
for idx, row in df.iterrows():
    if pd.notna(row.get("dopant")) and str(row.get("dopant", "")).strip():
        continue
    combined = " ".join([str(row.get("title", "") or ""), str(row.get("abstract", "") or "")])
    for d, pat in zip(DOPANTS, DOPANT_PATTERNS):
        if pat.search(combined):
            df.at[idx, "dopant"] = d
            filled_dopant += 1
            break
print(f"  도핑원소 추가:       {filled_dopant:,}편")

# ── 5. 결정상(crystal_phase) 키워드 ─────────────────────────────────────────
PHASE_KEYWORDS = [
    ("fluorite cubic", ["fluorite", "cubic fluorite", "face-centered cubic"]),
    ("Ce2O3",          ["ce2o3", "cerium(iii) oxide", "cerium sesquioxide"]),
    ("mixed",          ["mixed phase", "mixed oxide", "multiphase"]),
]
filled_phase = 0
for idx, row in df.iterrows():
    if pd.notna(row.get("crystal_phase")) and str(row.get("crystal_phase", "")).strip():
        continue
    combined = " ".join([str(row.get("title", "") or ""), str(row.get("abstract", "") or "")]).lower()
    for phase, keywords in PHASE_KEYWORDS:
        if any(kw in combined for kw in keywords):
            df.at[idx, "crystal_phase"] = phase
            filled_phase += 1
            break
print(f"  결정상 추가:         {filled_phase:,}편")

# ── 결과 저장 ────────────────────────────────────────────────────────────────
# 기존 시트 유지, 합성조건 시트만 교체
existing_sheets = {}
xl = pd.ExcelFile(_output_path)
for sheet in xl.sheet_names:
    if sheet != "합성조건":
        existing_sheets[sheet] = xl.parse(sheet)
xl.close()

with pd.ExcelWriter(_output_path, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="합성조건", index=False)
    for sheet_name, sheet_df in existing_sheets.items():
        sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)

print(f"\n저장 완료: {_output_path}")
print(f"\n── 최종 채움률 ──────────────────────────────────────")
key_cols = ["synthesis_method", "morphology", "crystallite_size_xrd_nm",
            "particle_size_tem_nm", "bet_surface_area", "dopant", "crystal_phase"]
for col in key_cols:
    if col in df.columns:
        n = df[col].replace("", pd.NA).notna().sum()
        print(f"  {col:<32} {n:>5,}  ({n/total*100:.1f}%)")


# %% [셀 18] LLM 재추출 - 개선된 프롬프트 + 프롬프트 캐싱으로 누락 필드 보완
"""
전문(full text)이 있으나 핵심 합성 필드가 비어있는 논문을 대상으로
개선된 LLM 프롬프트로 재추출합니다.

개선 사항:
  1. 합성방법 추론 규칙 명시 (autoclave → hydrothermal 등)
  2. Ce 전구체 정식명 → 화학식 매핑 힌트 제공
  3. Anthropic 프롬프트 캐싱으로 비용 약 70% 절감
  4. 기존 캐시의 non-null 값은 덮어쓰지 않고 빈 필드만 보완
"""
import os, json, re, time
import pandas as pd
from dotenv import load_dotenv
from tqdm import tqdm

_BASE_DIR    = r"d:\머신러닝 교육\ceria_pipeline_data"
_CACHE_PATH  = os.path.join(_BASE_DIR, "output", "llm_cache.json")
_OUTPUT_PATH = os.path.join(_BASE_DIR, "output", "ceria_synthesis_database.xlsx")

load_dotenv(os.path.join(_BASE_DIR, ".env"))
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
OPENAI_API_KEY    = os.environ.get("OPENAI_API_KEY", "")

if not ANTHROPIC_API_KEY and not OPENAI_API_KEY:
    raise SystemExit("API 키 없음 → .env에 ANTHROPIC_API_KEY 또는 OPENAI_API_KEY 설정 후 실행")

# ── 기존 캐시 로드 ─────────────────────────────────────────────────────────────
with open(_CACHE_PATH, encoding="utf-8") as _f:
    _existing = json.load(_f)
cache_map = {r.get("paper_id"): r for r in _existing}
print(f"기존 캐시: {len(cache_map)}편")

df18 = pd.read_excel(_OUTPUT_PATH, sheet_name="합성조건")

# ── 재추출 대상 선별 ───────────────────────────────────────────────────────────
_KEY_FIELDS = ["ce_precursor", "synthesis_method", "synthesis_temperature_c"]
_SYNTH_KW   = [
    "ce(no3)", "cecl", "cerium nitrate", "cerium chloride", "cerium acetate",
    "hydrothermal", "autoclave", "sol-gel", "precipitation", "coprecipitation",
    "combustion", "calcin", "°c", "dissolv", "solvothermal", "reflux",
]

def _empty(v):
    return pd.isna(v) or str(v).strip().lower() in ("", "nan", "none")

_targets = []
for _, _row in df18.iterrows():
    _tp = str(_row.get("text_path", "") or "")
    if not _tp or not os.path.exists(_tp):
        continue
    if not any(_empty(_row.get(f)) for f in _KEY_FIELDS):
        continue
    try:
        with open(_tp, encoding="utf-8", errors="ignore") as _f:
            _sample = _f.read(4000).lower()
        if not any(kw in _sample for kw in _SYNTH_KW):
            continue
    except Exception:
        continue
    _targets.append({
        "paper_id":      _row["paper_id"],
        "title":         str(_row.get("title", "") or ""),
        "text_path":     _tp,
        "missing_fields": [f for f in _KEY_FIELDS if _empty(_row.get(f))],
    })

_cost_unit = 0.0003 if OPENAI_API_KEY else 0.00025
print(f"재추출 대상: {len(_targets)}편")
print(f"예상 비용:  ~${len(_targets) * _cost_unit:.2f} "
      f"({'OpenAI gpt-4o-mini' if OPENAI_API_KEY else 'Anthropic claude-haiku + 프롬프트 캐싱'})")

# ── 개선된 시스템 프롬프트 ─────────────────────────────────────────────────────
_SYSTEM_PROMPT = """You are an expert materials scientist specializing in CeO2 (ceria) nanoparticle synthesis. Extract synthesis conditions from paper text and return ONLY a valid JSON object. Use null only when information is truly absent.

=== SYNTHESIS METHOD INFERENCE (infer even if not explicitly named) ===
- "autoclave", "Teflon-lined vessel", "sealed vessel" → hydrothermal
- "reflux" + organic solvent → solvothermal
- "co-precipitation", "NaOH/NH3 added dropwise", "pH adjustment" + precipitation → co-precipitation
- "NaOH/NH3" + no autoclave → precipitation
- "citric acid" + heating to gel, "Pechini", "sol gel", "alkoxide" → sol-gel
- "glycine fuel", "urea fuel", "self-ignition", "self-combustion" → combustion
- "microwave irradiation/treatment" → microwave
- "ball mill", "mechanical grinding" → mechanochemical
- "sonication" during synthesis step → sonochemical
- "hard template", "soft template", "SBA-15", "P123" → template
- "thermal decomposition", "pyrolysis of precursor" → thermal_decomposition
When uncertain, use "wet chemical" rather than null.

=== Ce PRECURSOR → USE CANONICAL FORMULA ===
cerium(III) nitrate hexahydrate / cerous nitrate / Ce(NO3)3 → Ce(NO3)3·6H2O
cerium(III) chloride / CeCl3 → CeCl3·7H2O
cerium(III) acetate → Ce(CH3COO)3
ammonium cerium(IV) nitrate / ceric ammonium nitrate / CAN → (NH4)2Ce(NO3)6
cerium(III) acetylacetonate / Ce(acac)3 → Ce(acac)3
cerium(IV) sulfate / ceric sulfate → Ce(SO4)2
cerium isopropoxide → Ce(OiPr)4
cerium ethoxide → Ce(OEt)4
cerium carbonate → Ce2(CO3)3
cerium oxalate → Ce2(C2O4)3

=== OUTPUT JSON FIELDS (all required, null if absent) ===
synthesis_method, ce_precursor, ce_precursor_amount, solvent, solvent_amount,
additive, additive_amount, synthesis_temperature_c (NUMBER °C),
synthesis_time_h (NUMBER hours; overnight=12, 1 day=24),
atmosphere (air/N2/Ar/O2/vacuum/autoclave/other),
calcination_temperature_c (NUMBER °C), calcination_time_h (NUMBER h),
drying_temperature_c (NUMBER °C),
particle_size_tem_nm (NUMBER, TEM only), particle_size_sem_nm (NUMBER, SEM only),
crystallite_size_xrd_nm (NUMBER, XRD Scherrer only),
morphology (sphere/cube/rod/wire/flower/octahedron/plate/porous/hollow/other),
crystal_phase (e.g. "fluorite cubic"), dopant (element symbol, null if pure CeO2),
bet_surface_area (NUMBER m²/g), notes (one sentence)"""

# ── 실험 섹션 추출 (개선) ──────────────────────────────────────────────────────
_SEC_START = re.compile(
    r'\n[ \t]*(?:\d+[\.\d]*\s*)?'
    r'(?:experimental(?:\s+(?:section|details?|procedure|part|methods?))?'
    r'|materials?\s+and\s+(?:experimental\s+)?methods?'
    r'|synthesis(?:\s+(?:and\s+(?:characterization|fabrication)|of\s+[\w\s]{0,35}))?'
    r'|preparation\s+of\s+[\w\s]{0,45}'
    r'|sample\s+preparation|fabrication(?:\s+of\s+[\w\s]{0,30})?'
    r'|catalyst\s+preparation|nanoparticle\s+synthesis)'
    r'\s*[\n:]',
    re.IGNORECASE,
)
_SEC_END = re.compile(
    r'\n[ \t]*(?:\d+[\.\d]*\s*)?'
    r'(?:results?(?:\s+and\s+discussion)?|discussion|characterization'
    r'|conclusions?|references?|acknowledgements?|bibliography'
    r'|supporting\s+information)\s*\n',
    re.IGNORECASE,
)

def _get_section(text: str, max_chars: int = 7000) -> str:
    m = _SEC_START.search(text)
    if m:
        start = m.end()
        em = _SEC_END.search(text, start)
        snippet = text[start:(em.start() if em else start + max_chars)].strip()
        if len(snippet) > 400:
            return snippet[:max_chars]
    return text[:max_chars]

# ── LLM 클라이언트 초기화 ──────────────────────────────────────────────────────
if OPENAI_API_KEY:
    from openai import OpenAI as _OAI
    _oai = _OAI(api_key=OPENAI_API_KEY)

    def _call_llm(snippet: str, pid, title: str):
        try:
            r = _oai.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": _SYSTEM_PROMPT},
                    {"role": "user",   "content": f"Title: {title}\n\n{snippet}"},
                ],
                max_tokens=1400, temperature=0,
                response_format={"type": "json_object"},
            )
            return _parse_llm(r.choices[0].message.content, pid, title)
        except Exception as e:
            print(f"  OpenAI 오류 ({str(pid)[:20]}): {e}")
            return None

else:
    import anthropic as _ant
    _ant_client = _ant.Anthropic(api_key=ANTHROPIC_API_KEY)

    def _call_llm(snippet: str, pid, title: str):
        try:
            # system prompt에 캐시 제어 적용 → 반복 호출 시 토큰 비용 ~70% 절감
            r = _ant_client.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=1400,
                system=[{
                    "type": "text",
                    "text": _SYSTEM_PROMPT,
                    "cache_control": {"type": "ephemeral"},
                }],
                messages=[{
                    "role": "user",
                    "content": f"Title: {title}\n\n{snippet}",
                }],
            )
            return _parse_llm(r.content[0].text, pid, title)
        except Exception as e:
            print(f"  Claude 오류 ({str(pid)[:20]}): {e}")
            return None

def _parse_llm(content: str, pid, title: str):
    m = re.search(r"\{.*\}", content, re.DOTALL)
    if m:
        try:
            rec = json.loads(m.group())
            rec["paper_id"] = pid
            rec["title"]    = title
            return rec
        except json.JSONDecodeError:
            pass
    return None

# ── 재추출 실행 ────────────────────────────────────────────────────────────────
_delay = 0.1 if OPENAI_API_KEY else 0.25
_updated = _added = 0

for _item in tqdm(_targets, desc="LLM 재추출"):
    _pid   = _item["paper_id"]
    _title = _item["title"]
    try:
        with open(_item["text_path"], encoding="utf-8", errors="ignore") as _f:
            _full = _f.read()
    except Exception:
        continue

    _result = _call_llm(_get_section(_full), _pid, _title)
    if not _result:
        continue

    if _pid in cache_map:
        _old = cache_map[_pid]
        for _k, _v in _result.items():
            if _v is not None and _empty(_old.get(_k)):
                _old[_k] = _v
        _updated += 1
    else:
        cache_map[_pid] = _result
        _added += 1

    time.sleep(_delay)

print(f"\n기존 레코드 보완: {_updated}편 | 신규 추가: {_added}편")

# ── 캐시 저장 ─────────────────────────────────────────────────────────────────
_updated_cache = list(cache_map.values())
with open(_CACHE_PATH, "w", encoding="utf-8") as _f:
    json.dump(_updated_cache, _f, ensure_ascii=False)
print(f"캐시 저장 완료: {len(_updated_cache)}편 → {_CACHE_PATH}")
print("\n→ 셀 14를 실행하면 업데이트된 캐시로 Excel이 재생성됩니다.")


# %% [셀 19] 종합 후처리 - 신규 필드 추출 + 데이터 검증 + 완성도 점수
"""
기존 Excel에서 단독 실행 가능한 종합 후처리 셀.

수행 작업:
  1. 도핑 농도(dopant_concentration) / Ce-도펀트 화학식(dopant_formula) 추출
  2. 합성 pH(ph_synthesis) 추출
  3. 전구체 몰농도(precursor_concentration) 추출
  4. 데이터 검증 — 물리적 범위 이탈 값 플래그(data_quality_flags)
  5. 완성도 점수(completeness_score, 0~100%) 산출
  6. 합성 논문 여부 분류(is_synthesis_paper)
  7. 업데이트된 Excel 저장
"""
import pandas as pd, os, re

_BASE = r"d:\머신러닝 교육\ceria_pipeline_data"
_PATH = os.path.join(_BASE, "output", "ceria_synthesis_database.xlsx")
df19 = pd.read_excel(_PATH, sheet_name="합성조건")
total19 = len(df19)
print(f"로드: {total19:,}편\n")

# ── 헬퍼 ──────────────────────────────────────────────────────────────────────
def _empty19(v):
    return pd.isna(v) or str(v).strip().lower() in ("", "nan", "none")

def _text19(row):
    return " ".join([str(row.get("title","") or ""), str(row.get("abstract","") or "")])

# ─────────────────────────────────────────────────────────────────────────────
# 1. 도핑 농도(dopant_concentration) & Ce-도펀트 화학식(dopant_formula)
# ─────────────────────────────────────────────────────────────────────────────
_CONC_MOL = re.compile(r'(\d+(?:\.\d+)?)\s*(?:mol|at|mole|atomic)\s*%', re.I)
_CONC_X   = re.compile(r'\bx\s*=\s*(\d*\.?\d+)', re.I)
_FORMULA  = re.compile(
    r'Ce[_\s]?(\d*\.?\d+)\s*([A-Z][a-z]?)[_\s]?(\d*\.?\d+)\s*O[_\s]?\d*(?:[_\s]?-?\s*[δd])?',
    re.I
)

if "dopant_concentration" not in df19.columns:
    df19["dopant_concentration"] = None
if "dopant_formula" not in df19.columns:
    df19["dopant_formula"] = None

cnt_dc = cnt_df = 0
for idx, row in df19.iterrows():
    txt = _text19(row)

    # 화학식 먼저 (Ce0.9Sm0.1O2 등)
    if _empty19(row.get("dopant_formula")):
        mf = _FORMULA.search(txt)
        if mf:
            df19.at[idx, "dopant_formula"] = mf.group(0).strip()
            cnt_df += 1
            # 화학식에서 농도 추출
            if _empty19(row.get("dopant_concentration")):
                try:
                    frac = float(mf.group(3))
                    if 0 < frac < 1:
                        df19.at[idx, "dopant_concentration"] = f"{frac*100:.1f} mol%"
                        cnt_dc += 1
                except (ValueError, IndexError):
                    pass

    # mol% / at% 직접 표기
    if _empty19(row.get("dopant_concentration")):
        mc = _CONC_MOL.search(txt)
        if mc:
            try:
                v = float(mc.group(1))
                if 0 < v <= 100:
                    df19.at[idx, "dopant_concentration"] = f"{v} mol%"
                    cnt_dc += 1
            except (ValueError, IndexError):
                pass

    # x = 0.1 표기
    if _empty19(row.get("dopant_concentration")):
        mx = _CONC_X.search(txt)
        if mx:
            try:
                v = float(mx.group(1))
                if 0 < v < 1:
                    df19.at[idx, "dopant_concentration"] = f"x={v}"
                    cnt_dc += 1
            except (ValueError, IndexError):
                pass

print(f"  dopant_formula      추가: {cnt_df:,}편")
print(f"  dopant_concentration 추가: {cnt_dc:,}편")

# ─────────────────────────────────────────────────────────────────────────────
# 2. 합성 pH (ph_synthesis)
# ─────────────────────────────────────────────────────────────────────────────
_PH_RE = [
    re.compile(r'pH\s+(?:was\s+)?(?:adjusted\s+to\s+|set\s+to\s+|of\s+|=\s*|≈\s*)?(\d+(?:\.\d+)?)', re.I),
    re.compile(r'pH\s*[=:≈]\s*(\d+(?:\.\d+)?)', re.I),
    re.compile(r'(?:to|at)\s+pH\s+(\d+(?:\.\d+)?)', re.I),
]

if "ph_synthesis" not in df19.columns:
    df19["ph_synthesis"] = None

cnt_ph = 0
for idx, row in df19.iterrows():
    if not _empty19(row.get("ph_synthesis")):
        continue
    txt = _text19(row)
    for pat in _PH_RE:
        m = pat.search(txt)
        if m:
            try:
                v = float(m.group(1))
                if 0 <= v <= 14:
                    df19.at[idx, "ph_synthesis"] = v
                    cnt_ph += 1
                    break
            except (ValueError, IndexError):
                pass
print(f"  ph_synthesis        추가: {cnt_ph:,}편")

# ─────────────────────────────────────────────────────────────────────────────
# 3. 전구체 몰농도 (precursor_concentration)
# ─────────────────────────────────────────────────────────────────────────────
_PREC_CONC = [
    re.compile(r'(\d+(?:\.\d+)?)\s*(?:M|mol/L|mmol/L|mM)\b', re.I),
    re.compile(r'(\d+(?:\.\d+)?)\s*mol(?:ar)?\s+(?:solution|concentration)', re.I),
]

if "precursor_concentration" not in df19.columns:
    df19["precursor_concentration"] = None

cnt_pc = 0
for idx, row in df19.iterrows():
    if not _empty19(row.get("precursor_concentration")):
        continue
    txt = _text19(row)
    for pat in _PREC_CONC:
        m = pat.search(txt)
        if m:
            df19.at[idx, "precursor_concentration"] = m.group(0).strip()
            cnt_pc += 1
            break
print(f"  precursor_concentration 추가: {cnt_pc:,}편")

# ─────────────────────────────────────────────────────────────────────────────
# 4. 데이터 검증 — 물리적 범위 이탈 플래그 (data_quality_flags)
# ─────────────────────────────────────────────────────────────────────────────
VALID_RANGES = {
    "synthesis_temperature_c":    (0,   1500),
    "calcination_temperature_c":  (50,  1600),
    "drying_temperature_c":       (20,  500),
    "synthesis_time_h":           (0.01, 1000),
    "calcination_time_h":         (0.01, 200),
    "particle_size_tem_nm":       (0.3,  1000),
    "particle_size_sem_nm":       (0.3,  5000),
    "crystallite_size_xrd_nm":    (0.3,  500),
    "bet_surface_area":           (0.1,  1500),
    "ph_synthesis":               (0,    14),
}

df19["data_quality_flags"] = ""
cnt_flags = 0
for idx, row in df19.iterrows():
    flags = []
    for col, (lo, hi) in VALID_RANGES.items():
        v = row.get(col)
        if _empty19(v):
            continue
        try:
            fv = float(v)
            if not (lo <= fv <= hi):
                flags.append(f"{col}={fv}(범위:{lo}-{hi})")
        except (ValueError, TypeError):
            pass
    if flags:
        df19.at[idx, "data_quality_flags"] = "; ".join(flags)
        cnt_flags += 1

print(f"  data_quality_flags  이상값: {cnt_flags:,}편")

# ─────────────────────────────────────────────────────────────────────────────
# 5. 완성도 점수 (completeness_score, 0~100%)
# ─────────────────────────────────────────────────────────────────────────────
# 가중치: 핵심 필드는 2.0, 중요 필드는 1.5, 일반 필드는 1.0
FIELD_WEIGHTS = {
    "synthesis_method":         2.0,
    "ce_precursor":             2.0,
    "synthesis_temperature_c":  1.5,
    "synthesis_time_h":         1.5,
    "solvent":                  1.5,
    "additive":                 1.0,
    "atmosphere":               1.0,
    "calcination_temperature_c": 1.0,
    "particle_size_tem_nm":     1.5,
    "crystallite_size_xrd_nm":  1.5,
    "particle_size_sem_nm":     1.0,
    "morphology":               1.5,
    "crystal_phase":            1.0,
    "dopant":                   1.0,
    "dopant_concentration":     1.0,
    "bet_surface_area":         1.0,
    "ph_synthesis":             1.0,
}
MAX_SCORE = sum(FIELD_WEIGHTS.values())

df19["completeness_score"] = df19.apply(
    lambda row: round(
        sum(w for f, w in FIELD_WEIGHTS.items() if not _empty19(row.get(f)))
        / MAX_SCORE * 100, 1
    ), axis=1
)
score_mean = df19["completeness_score"].mean()
score_ge50 = (df19["completeness_score"] >= 50).sum()
print(f"  completeness_score  평균: {score_mean:.1f}% | 50%이상: {score_ge50:,}편")

# ─────────────────────────────────────────────────────────────────────────────
# 6. 합성 논문 여부 분류 (is_synthesis_paper)
# ─────────────────────────────────────────────────────────────────────────────
_SYNTH_PAPER_KW = [
    "synthesized", "prepared", "fabricated", "synthesis", "preparation",
    "hydrothermal", "sol-gel", "precipitation", "calcin", "autoclave",
    "ce(no3)", "cecl", "cerium nitrate", "cerium chloride",
]
_NON_SYNTH_KW = [
    "review", "theoretical", "dft calculation", "molecular dynamics",
    "computational", "simulation", "first-principles",
]

df19["is_synthesis_paper"] = df19.apply(
    lambda row: (
        any(kw in _text19(row).lower() for kw in _SYNTH_PAPER_KW) and
        not any(kw in _text19(row).lower() for kw in _NON_SYNTH_KW)
    ), axis=1
)
synth_count = df19["is_synthesis_paper"].sum()
print(f"  is_synthesis_paper  합성논문: {synth_count:,}편 / {total19:,}편")

# ─────────────────────────────────────────────────────────────────────────────
# 7. 저장 (기존 시트 유지, 합성조건 시트 교체)
# ─────────────────────────────────────────────────────────────────────────────
_xl = pd.ExcelFile(_PATH)
_sheets = {s: _xl.parse(s) for s in _xl.sheet_names if s != "합성조건"}
_xl.close()

with pd.ExcelWriter(_PATH, engine="openpyxl") as _w:
    df19.to_excel(_w, sheet_name="합성조건", index=False)
    for _sn, _sd in _sheets.items():
        _sd.to_excel(_w, sheet_name=_sn, index=False)

print(f"\n저장 완료: {_PATH}")
print(f"\n── 최종 채움률 (신규 포함) ─────────────────────────────")
for col in ["synthesis_method", "ce_precursor", "dopant_concentration",
            "dopant_formula", "ph_synthesis", "precursor_concentration",
            "completeness_score"]:
    if col in df19.columns:
        if col == "completeness_score":
            print(f"  {col:<32} 평균 {df19[col].mean():.1f}%")
        else:
            n = df19[col].replace("", pd.NA).notna().sum()
            print(f"  {col:<32} {n:>5,}  ({n/total19*100:.1f}%)")


# %% [셀 20] Anthropic Batch API - 대규모 재추출 (비용 50% 절감)
"""
셀 18은 동기 API를 사용합니다.
이 셀은 Anthropic Message Batches API를 사용하여 최대 10,000건을 비동기로 처리합니다.
비용: 동기 대비 50% 절감 (배치 할인)

실행 흐름:
  1. 배치 생성 및 제출 → batch_id 출력
  2. 배치 완료 대기 (polling)
  3. 결과 수집 → llm_cache 업데이트
  4. 셀 14 실행하면 Excel 재생성
"""
import os, json, re, time
import pandas as pd
from dotenv import load_dotenv

_BASE = r"d:\머신러닝 교육\ceria_pipeline_data"
load_dotenv(os.path.join(_BASE, ".env"))
_ANT_KEY = os.environ.get("ANTHROPIC_API_KEY", "")

if not _ANT_KEY:
    raise SystemExit("ANTHROPIC_API_KEY 필요. OpenAI는 Batch API 미지원(별도 구현 필요)")

import anthropic as _ant
_client20 = _ant.Anthropic(api_key=_ANT_KEY)

_CACHE  = os.path.join(_BASE, "output", "llm_cache.json")
_OUTXLS = os.path.join(_BASE, "output", "ceria_synthesis_database.xlsx")

_df20 = pd.read_excel(_OUTXLS, sheet_name="합성조건")
_KEY20 = ["ce_precursor", "synthesis_method", "synthesis_temperature_c",
          "dopant_concentration", "dopant_formula", "ph_synthesis"]

def _empty20(v):
    return pd.isna(v) or str(v).strip().lower() in ("", "nan", "none")

# ── 배치 대상 선별 ─────────────────────────────────────────────────────────────
_KW20 = ["ce(no3)", "cecl", "cerium nitrate", "cerium chloride",
         "hydrothermal", "autoclave", "sol-gel", "precipitation",
         "calcin", "°c", "solvothermal", "combustion"]

_batch_items = []
for _, _row in _df20.iterrows():
    _tp = str(_row.get("text_path","") or "")
    if not _tp or not os.path.exists(_tp):
        continue
    if not any(_empty20(_row.get(f)) for f in _KEY20):
        continue
    try:
        with open(_tp, encoding="utf-8", errors="ignore") as _f:
            _s = _f.read(3000).lower()
        if not any(k in _s for k in _KW20):
            continue
    except Exception:
        continue
    _batch_items.append({
        "paper_id": _row["paper_id"],
        "title":    str(_row.get("title","") or ""),
        "text_path": _tp,
    })

print(f"배치 대상: {len(_batch_items)}편")
print(f"예상 비용: ~${len(_batch_items) * 0.000125:.2f} (haiku 배치 할인 50%)")
if not _batch_items:
    raise SystemExit("재추출 대상 없음")

# ── 배치 요청 생성 ─────────────────────────────────────────────────────────────
_SEC_RE20 = re.compile(
    r'\n[ \t]*(?:\d+[\.\d]*\s*)?(?:experimental|materials?\s+and\s+methods?|synthesis|preparation)'
    r'[^a-z][^\n]*\n', re.IGNORECASE
)
_END_RE20 = re.compile(
    r'\n[ \t]*(?:\d+[\.\d]*\s*)?(?:results?|discussion|conclusion|references?)\s*\n',
    re.IGNORECASE
)

def _section20(text, max_chars=7000):
    m = _SEC_RE20.search(text)
    if m:
        s = m.end()
        e = _END_RE20.search(text, s)
        snip = text[s:(e.start() if e else s+max_chars)].strip()
        if len(snip) > 400:
            return snip[:max_chars]
    return text[:max_chars]

# SYSTEM_PROMPT은 셀 18의 _SYSTEM_PROMPT 재사용
try:
    _sys20 = _SYSTEM_PROMPT
except NameError:
    _sys20 = """You are an expert materials scientist. Extract CeO2 synthesis conditions. Return ONLY valid JSON with fields: synthesis_method, ce_precursor, ce_precursor_amount, precursor_concentration, solvent, solvent_amount, additive, additive_amount, synthesis_temperature_c, synthesis_time_h, ph_synthesis, atmosphere, calcination_temperature_c, calcination_time_h, drying_temperature_c, particle_size_tem_nm, particle_size_sem_nm, crystallite_size_xrd_nm, morphology, crystal_phase, dopant, dopant_concentration, dopant_formula, bet_surface_area, notes. Use null for missing values."""

_BATCH_SIZE = 1000  # 회당 최대 1000건 (API 제한)
_batches_submitted = []

for _chunk_start in range(0, len(_batch_items), _BATCH_SIZE):
    _chunk = _batch_items[_chunk_start:_chunk_start + _BATCH_SIZE]
    _requests = []

    for _item in _chunk:
        try:
            with open(_item["text_path"], encoding="utf-8", errors="ignore") as _f:
                _full = _f.read()
        except Exception:
            continue
        _snip = _section20(_full)
        _requests.append(
            _ant.types.MessageCreateParamsNonStreaming(
                model="claude-haiku-4-5-20251001",
                max_tokens=1400,
                system=[{
                    "type": "text",
                    "text": _sys20,
                    "cache_control": {"type": "ephemeral"},
                }],
                messages=[{
                    "role": "user",
                    "content": f"custom_id:{_item['paper_id']}\nTitle: {_item['title']}\n\n{_snip}",
                }],
            )
        )

    if not _requests:
        continue

    try:
        _batch = _client20.beta.messages.batches.create(requests=[
            _ant.types.MessageBatchRequest(
                custom_id=str(_item["paper_id"]),
                params=_ant.types.MessageCreateParamsNonStreaming(
                    model="claude-haiku-4-5-20251001",
                    max_tokens=1400,
                    system=[{"type":"text","text":_sys20,"cache_control":{"type":"ephemeral"}}],
                    messages=[{"role":"user","content":
                        f"Title: {_item['title']}\n\n{_section20(open(_item['text_path'],encoding='utf-8',errors='ignore').read())}"}],
                )
            )
            for _item in _chunk
            if os.path.exists(_item["text_path"])
        ])
        _batches_submitted.append(_batch.id)
        print(f"배치 제출: {_batch.id} ({len(_chunk)}건)")
    except Exception as e:
        print(f"배치 생성 오류: {e}")

if not _batches_submitted:
    raise SystemExit("배치 제출 실패")

print(f"\n제출된 배치 ID들: {_batches_submitted}")
print("완료까지 수분~수시간 소요. 아래 셀 20b를 실행하여 결과를 수집하세요.")
print("또는 https://console.anthropic.com 에서 배치 상태 확인 가능")

# 배치 ID 저장 (재실행 시 사용)
_bid_path = os.path.join(_BASE, "output", "batch_ids.json")
with open(_bid_path, "w") as _f:
    json.dump(_batches_submitted, _f)
print(f"배치 ID 저장: {_bid_path}")


# %% [셀 20b] Batch API 결과 수집
"""
셀 20에서 제출한 배치의 결과를 수집합니다.
배치가 완료되면 실행하세요 (보통 수분~1시간 후).
"""
import os, json, re, time
import pandas as pd

_BASE = r"d:\머신러닝 교육\ceria_pipeline_data"
_CACHE  = os.path.join(_BASE, "output", "llm_cache.json")
_BID_PATH = os.path.join(_BASE, "output", "batch_ids.json")

try:
    _ANT_KEY
except NameError:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(_BASE, ".env"))
    _ANT_KEY = os.environ.get("ANTHROPIC_API_KEY", "")

import anthropic as _ant
_client20b = _ant.Anthropic(api_key=_ANT_KEY)

with open(_BID_PATH) as _f:
    _batch_ids = json.load(_f)

with open(_CACHE, encoding="utf-8") as _f:
    _cache20b = json.load(_f)
_cmap20b = {r.get("paper_id"): r for r in _cache20b}

_total_collected = 0

for _bid in _batch_ids:
    _batch_status = _client20b.beta.messages.batches.retrieve(_bid)
    print(f"배치 {_bid}: {_batch_status.processing_status}")

    if _batch_status.processing_status != "ended":
        print("  아직 처리 중입니다. 나중에 다시 실행하세요.")
        continue

    # 결과 수집
    for _result in _client20b.beta.messages.batches.results(_bid):
        _pid = _result.custom_id
        if _result.result.type != "succeeded":
            continue
        _content = _result.result.message.content[0].text
        _m = re.search(r"\{.*\}", _content, re.DOTALL)
        if not _m:
            continue
        try:
            _rec = json.loads(_m.group())
            _rec["paper_id"] = _pid
            if _pid in _cmap20b:
                _old = _cmap20b[_pid]
                for _k, _v in _rec.items():
                    if _v is not None and (pd.isna(_old.get(_k,"")) or str(_old.get(_k,"")).strip() in ("","nan","none")):
                        _old[_k] = _v
            else:
                _cmap20b[_pid] = _rec
            _total_collected += 1
        except (json.JSONDecodeError, Exception):
            pass

print(f"\n수집 완료: {_total_collected}건")

with open(_CACHE, "w", encoding="utf-8") as _f:
    json.dump(list(_cmap20b.values()), _f, ensure_ascii=False)
print(f"캐시 업데이트: {len(_cmap20b)}편 → {_CACHE}")
print("→ 셀 14를 실행하여 Excel 재생성")

